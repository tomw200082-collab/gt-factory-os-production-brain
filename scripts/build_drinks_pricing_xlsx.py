#!/usr/bin/env python3
"""
GT Everyday — customer-facing drinks pricing sheet (81% repricing, 2026-08-26).

Every figure is read from .claude/skills/drinks-pricelist/drinks_final_figures.json,
the approved and frozen source. Nothing is retyped and nothing is recomputed for
display: margin and profit are printed exactly as approved. The script recomputes
them only as a self-check and refuses to write the file if any row disagrees.

This sheet goes to customers who ask to see the costing, so it deliberately carries
no GT wholesale price, no landed cost and no supplier name.

Output: docs/pricing/2026-08-26_drinks_pricing_81pct.xlsx
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
FIG = ROOT / ".claude/skills/drinks-pricelist/drinks_final_figures.json"
OUT = ROOT / "docs/pricing/2026-08-26_drinks_pricing_81pct.xlsx"

# (code, hebrew name, latin name, page numbers) — mirrors the catalog's own bands
FAMILIES = [
    ("01", "חליטות קרות", "iced tea", [8, 9, 10, 11, 12, 13, 14]),
    ("02", "לימונדות", "lemonade", [16, 17, 18]),
    ("03", "משקאות דגל", "signature", [20, 21, 22, 23]),
    ("04", "גזוז", "gazoz", [25, 26, 27]),
    ("05", "אייס מאצ'ה", "ice matcha", [29, 30, 31, 32, 33, 34]),
    ("06", "מאצ'ה ספיישלס", "matcha specials", [36, 37, 38, 39, 40]),
    ("07", "מאצ'ה קוקוס", "matcha coconut", [42, 43, 44, 45, 46]),
    ("08", "צ'אי מסאלה", "chai massala", [48, 49, 50, 51, 52, 53]),
    ("09", "צ'אי קולד פואם", "chai cold foam", [55, 56, 57, 58]),
    ("10", "אייס אובה", "ice ube", [60, 61, 62, 63, 64]),
]

INK = "26221A"
BAND = "123B39"
ACCENT = "CC7A2E"
PAPER = "FBF8F2"

THIN = Side(style="thin", color="D8D2C4")


def num(s):
    """'₪4.73*' -> 4.73 ; '82%' -> 82 ; '₪13.84 לכוס' -> 13.84"""
    body = s.replace("₪", "").replace("%", "").replace("לכוס", "").strip()
    return float(body.rstrip("*"))


def main():
    fig = json.loads(FIG.read_text(encoding="utf-8"))
    meta, pages = fig["_meta"], fig["pages"]
    vat = 1.0 + meta["vat_rate"]
    cost_label, price_label = meta["labels"]["food_cost"], meta["labels"]["price"]

    mapped = sorted(p for *_, ps in FAMILIES for p in ps)
    if mapped != sorted(int(k) for k in pages):
        raise SystemExit("family map out of sync with the figures file")

    failures = []
    for p in mapped:
        d = pages[str(p)]
        cost, price = num(d["cost"]), num(d["price"])
        net = price / vat
        if price != int(price):
            failures.append(f"page {p}: price {d['price']} is not a whole shekel")
        if round((net - cost) / net * 100) != int(num(d["marg"])):
            failures.append(f"page {p}: margin {d['marg']} does not re-derive")
        if round(net - cost, 2) != num(d["prof"]):
            failures.append(f"page {p}: profit {d['prof']} does not re-derive")
    if failures:
        raise SystemExit("re-derivation failed:\n  " + "\n  ".join(failures))

    wb = Workbook()
    ws = wb.active
    ws.title = "מחירון משקאות"
    ws.sheet_view.rightToLeft = True

    ws["A1"] = "GT Everyday · מחירון משקאות מומלץ"
    ws["A1"].font = Font(size=16, bold=True, color=BAND)
    ws["A2"] = f"תקף מתאריך {meta['date']} · מחליף את {meta['supersedes']}"
    ws["A2"].font = Font(size=10, color="6B6455")
    ws["A3"] = "הרווח והמרווח מחושבים על המחיר ללא מע״מ (מחיר מומלץ ÷ 1.18)"
    ws["A3"].font = Font(size=10, color="6B6455")

    headers = [
        "עמוד",
        "משפחה",
        "משקה",
        cost_label,
        price_label,
        "מרווח",
        "רווח לכוס · ללא מע״מ",
    ]
    hrow = 5
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = Font(bold=True, color=PAPER, size=11)
        cell.fill = PatternFill("solid", fgColor=BAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hrow].height = 34

    r = hrow + 1
    for code, he, latin, page_nums in FAMILIES:
        band = ws.cell(row=r, column=1, value=f"{code} · {he} · {latin}")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        band.font = Font(bold=True, color=BAND, size=11)
        band.fill = PatternFill("solid", fgColor="EFE9DA")
        band.alignment = Alignment(horizontal="right", vertical="center")
        r += 1

        for p in page_nums:
            d = pages[str(p)]
            ws.cell(row=r, column=1, value=p).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=2, value=he)
            ws.cell(row=r, column=3, value=d["name"])
            ws.cell(row=r, column=4, value=d["cost"]).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=5, value=d["price"]).alignment = Alignment(horizontal="center")
            m = ws.cell(row=r, column=6, value=d["marg"])
            m.alignment = Alignment(horizontal="center")
            m.font = Font(bold=True, color=ACCENT)
            ws.cell(row=r, column=7, value=d["prof"]).alignment = Alignment(horizontal="center")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = Border(bottom=THIN)
            r += 1

    r += 1
    ws.cell(row=r, column=1, value="* כולל הערכת עלות גרניש/קצף").font = Font(size=9, color="6B6455")
    r += 1
    ws.cell(
        row=r,
        column=1,
        value="המחיר המומלץ הוא מחיר לצרכן כולל מע״מ; ה-FOOD COST הוא ללא מע״מ.",
    ).font = Font(size=9, color="6B6455")

    for c, w in enumerate([8, 20, 34, 22, 24, 10, 22], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)} — {len(mapped)} drinks, {len(FAMILIES)} families, 0 re-derivation failures")


if __name__ == "__main__":
    main()
