"""
Add FG-MAT-KIT (Matcha Accessory Kit) to GT_Master_Data.xlsx.

Changes:
  1. SUPPLIER_MASTER: append SUP-047 = Si Madeh / סי מדה
  2. COMPONENT_MASTER: append 6 new components
  3. ITEM_MASTER: append FG-MAT-KIT (REPACK, KIT)
  4. BOM_HEADER: append BOM-REPACK-MAT-KIT
  5. BOM_LINES: append 6 lines
  6. SUPPLIER_ITEM_MAP: append 2 known supplier mappings (Arizot 2100 → bottle, Si Madeh → 600ml cup)
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import load_workbook

XLSX = r"c:\Users\tomw2\GTeveryday Dropbox\Data Center\Tom\AI Agents & Projects\Code Agents\PRODUCTION\GT_Master_Data.xlsx"


def first_empty_row_after(ws, header_row_idx_1based, last_data_idx_0based):
    """Return 1-based row index of first empty row to append into, given last contiguous data row 0-based."""
    return last_data_idx_0based + 2  # convert to 1-based and move 1 down


def find_last_contiguous_data(ws, start_row_0based):
    rows = list(ws.iter_rows(values_only=True))
    last = start_row_0based - 1
    for i in range(start_row_0based, len(rows)):
        if rows[i][0] is not None:
            last = i
        else:
            break
    return last


wb = load_workbook(XLSX)

# ---------------------------------------------------------------------------
# 1. SUPPLIER_MASTER: append SUP-047
# ---------------------------------------------------------------------------
ws = wb["SUPPLIER_MASTER"]
# Find true last row with content (data is scattered: rows 5-29, then 33,35,36,38,39,40, then 999-1010).
last_row_1based = 0
for row in ws.iter_rows():
    if row[0].value is not None:
        last_row_1based = row[0].row
new_row = last_row_1based + 1
sup_row = [
    "SUP-047",                                  # SUPPLIER_ID
    "Si Madeh",                                 # SUPPLIER_NAME_OFFICIAL
    "סי מדה",                                   # SUPPLIER_NAME_SHORT
    "ACTIVE",                                   # STATUS
    "ACCESSORY",                                # SUPPLIER_TYPE
    None,                                       # PRIMARY_CONTACT_NAME
    None,                                       # PRIMARY_CONTACT_PHONE
    "ILS",                                      # CURRENCY
    None,                                       # PAYMENT_TERMS
    None,                                       # DEFAULT_LEAD_TIME_DAYS
    None,                                       # DEFAULT_MOQ
    "PENDING",                                  # APPROVAL_STATUS
    "Added 2026-05-07 for Matcha Kit (FG-MAT-KIT) — supplies 600ml matcha cup.",
]
for col_idx, val in enumerate(sup_row, start=1):
    ws.cell(row=new_row, column=col_idx, value=val)
print(f"SUPPLIER_MASTER: appended SUP-047 at row {new_row}")

# ---------------------------------------------------------------------------
# 2. COMPONENT_MASTER: append 6 components after row 146 (1-based)
# ---------------------------------------------------------------------------
ws = wb["COMPONENT_MASTER"]
last_data_0 = find_last_contiguous_data(ws, 5)
start_row = last_data_0 + 2  # 1-based first empty
print(f"COMPONENT_MASTER: appending starting row {start_row}")

# Columns:
# COMPONENT_ID, COMPONENT_NAME, COMPONENT_CLASS, COMPONENT_GROUP, STATUS,
# INVENTORY_UOM, PURCHASE_UOM, BOM_UOM, PURCHASE_TO_INV_FACTOR, PLANNING_POLICY,
# PRIMARY_SUPPLIER_ID, PRIMARY_SUPPLIER_NAME, ALTERNATE_SUPPLIERS, LEAD_TIME_DAYS,
# MOQ_PURCHASE_UOM, ORDER_MULTIPLE_PURCHASE_UOM, STD_COST_PER_PURCHASE_UOM,
# STD_COST_PER_INV_UOM, CRITICALITY, PLANNED_FLAG
new_components = [
    [
        "PKG-BOTTLE-MAT-KIT-500ML", "Bottle 500ML for Matcha Kit", "PACKAGING", "BOTTLE", "ACTIVE",
        "UNIT", "UNIT", "UNIT", 1, "LOT_FOR_LOT",
        "SUP-002", "Arizot 2100", None, None,
        None, None, None, None, "MEDIUM", "YES",
    ],
    [
        "PKG-CAP-MAT-KIT", "Cap dedicated for Matcha Kit", "PACKAGING", "CAP", "ACTIVE",
        "UNIT", "UNIT", "UNIT", 1, "LOT_FOR_LOT",
        None, None, None, None,
        None, None, None, None, "MEDIUM", "YES",
    ],
    [
        "PKG-CARTON-MAT-KIT", "GT-branded black carton for Matcha Kit", "PACKAGING", "CARTON", "ACTIVE",
        "UNIT", "UNIT", "UNIT", 1, "LOT_FOR_LOT",
        None, None, None, None,
        None, None, None, None, "MEDIUM", "YES",
    ],
    [
        "ACC-FROTHER-MAT", "Manual Matcha Frother", "ACCESSORY", "KIT_ACCESSORY", "ACTIVE",
        "UNIT", "UNIT", "UNIT", 1, "LOT_FOR_LOT",
        None, None, None, None,
        None, None, None, None, "MEDIUM", "YES",
    ],
    [
        "ACC-CUP-MAT-600ML", "Matcha Cup 600ML", "ACCESSORY", "KIT_ACCESSORY", "ACTIVE",
        "UNIT", "UNIT", "UNIT", 1, "LOT_FOR_LOT",
        "SUP-047", "Si Madeh", None, None,
        None, None, None, None, "MEDIUM", "YES",
    ],
    [
        "ACC-CUP-MEASURE-MAT", "Matcha Measuring Cup", "ACCESSORY", "KIT_ACCESSORY", "ACTIVE",
        "UNIT", "UNIT", "UNIT", 1, "LOT_FOR_LOT",
        None, None, None, None,
        None, None, None, None, "MEDIUM", "YES",
    ],
]
for r_offset, row_vals in enumerate(new_components):
    for c_idx, val in enumerate(row_vals, start=1):
        ws.cell(row=start_row + r_offset, column=c_idx, value=val)
    print(f"  + row {start_row + r_offset}: {row_vals[0]}")

# ---------------------------------------------------------------------------
# 3. ITEM_MASTER: append FG-MAT-KIT
# ---------------------------------------------------------------------------
ws = wb["ITEM_MASTER"]
last_data_0 = find_last_contiguous_data(ws, 5)
start_row = last_data_0 + 2
print(f"ITEM_MASTER: appending starting row {start_row}")

# Columns:
# ITEM_ID, ITEM_NAME, FAMILY, PACK_SIZE, SALES_UOM, SWEETNESS, SUPPLY_METHOD, ITEM_TYPE,
# STATUS, BARCODE, LEGACY_SKU, SHELF_LIFE_DAYS, STORAGE, CASE_PACK,
# PRIMARY_BOM_ID, BASE_BOM_ID, BASE_FILL_QTY_PER_UNIT, NOTES, SUB_TYPE, PRODUCT_GROUP
item_row = [
    "FG-MAT-KIT",                  # ITEM_ID
    "MATCHA KIT",                  # ITEM_NAME
    "MATCHA",                      # FAMILY
    "KIT",                         # PACK_SIZE
    "UNIT",                        # SALES_UOM
    "REGULAR",                     # SWEETNESS
    "REPACK",                      # SUPPLY_METHOD
    "KIT",                         # ITEM_TYPE
    "ACTIVE",                      # STATUS
    None,                          # BARCODE
    None,                          # LEGACY_SKU
    None,                          # SHELF_LIFE_DAYS
    "Ambient",                     # STORAGE
    None,                          # CASE_PACK
    "BOM-REPACK-MAT-KIT",          # PRIMARY_BOM_ID
    None,                          # BASE_BOM_ID
    0,                             # BASE_FILL_QTY_PER_UNIT
    "Accessory bundle: 2x bottle 500ML + 2x dedicated cap + frother + 600ml cup + measuring cup + GT-branded black carton.",
    "Matcha Kit",                  # SUB_TYPE
    "GT MATCHA",                   # PRODUCT_GROUP
]
for c_idx, val in enumerate(item_row, start=1):
    ws.cell(row=start_row, column=c_idx, value=val)
print(f"  + row {start_row}: FG-MAT-KIT")

# ---------------------------------------------------------------------------
# 4. BOM_HEADER: append BOM-REPACK-MAT-KIT
# ---------------------------------------------------------------------------
ws = wb["BOM_HEADER"]
last_data_0 = find_last_contiguous_data(ws, 5)
start_row = last_data_0 + 2
print(f"BOM_HEADER: appending starting row {start_row}")

# Columns:
# BOM_ID, BOM_KIND, DISPLAY_FAMILY, SWEETNESS, PACK_SIZE, PARENT_REF_TYPE, PARENT_REF_ID,
# PARENT_NAME, LINKED_BASE_BOM_ID, FINAL_BOM_OUTPUT_QTY, FINAL_BOM_OUTPUT_UOM, VERSION,
# STATUS, REVIEW_FLAG, OWNER_NOTES, MIN_RUN_L, BUFFER_PCT
bom_header_row = [
    "BOM-REPACK-MAT-KIT", "REPACK", "MATCHA", "REGULAR", "KIT",
    "ITEM", "FG-MAT-KIT", "MATCHA KIT", None,
    1, "UNIT", "V1_INITIAL", "ACTIVE", "OK",
    "Multi-component accessory kit for matcha. Created 2026-05-07. Suppliers TBD for cap, frother, measuring cup, carton.",
    0, 0,
]
for c_idx, val in enumerate(bom_header_row, start=1):
    ws.cell(row=start_row, column=c_idx, value=val)
print(f"  + row {start_row}: BOM-REPACK-MAT-KIT")

# ---------------------------------------------------------------------------
# 5. BOM_LINES: append 6 lines
# ---------------------------------------------------------------------------
ws = wb["BOM_LINES"]
last_data_0 = find_last_contiguous_data(ws, 5)
# Note: BOM_LINES has a known gap at rows 386-387; find_last_contiguous_data stops at first gap.
# So scan for the *true* last row with content.
true_last = last_data_0
for row in ws.iter_rows():
    if row[0].value is not None:
        true_last = row[0].row - 1  # convert to 0-based
start_row = true_last + 2
print(f"BOM_LINES: appending starting row {start_row}")

# Columns:
# LINE_KEY, BOM_ID, LINE_NO, BOM_KIND, DISPLAY_FAMILY, SWEETNESS, PACK_SIZE,
# COMPONENT_REF_TYPE, FINAL_COMPONENT_ID, FINAL_COMPONENT_NAME, FINAL_COMPONENT_QTY,
# COMPONENT_UOM, STATUS, LINE_CHECK, OWNER_NOTES, STD_COST_PER_UOM, LINE_STD_COST,
# SCALING_METHOD, QTY_PER_L_OUTPUT
bom_lines = [
    ["BOM-REPACK-MAT-KIT-L01", "BOM-REPACK-MAT-KIT", 1, "REPACK", "MATCHA", "REGULAR", "KIT",
     "RAW_NAME", "PKG-BOTTLE-MAT-KIT-500ML", "Bottle 500ML for Matcha Kit", 2,
     "UNIT", "ACTIVE", "OK", None, None, None, "FIXED", None],
    ["BOM-REPACK-MAT-KIT-L02", "BOM-REPACK-MAT-KIT", 2, "REPACK", "MATCHA", "REGULAR", "KIT",
     "RAW_NAME", "PKG-CAP-MAT-KIT", "Cap dedicated for Matcha Kit", 2,
     "UNIT", "ACTIVE", "OK", None, None, None, "FIXED", None],
    ["BOM-REPACK-MAT-KIT-L03", "BOM-REPACK-MAT-KIT", 3, "REPACK", "MATCHA", "REGULAR", "KIT",
     "RAW_NAME", "ACC-FROTHER-MAT", "Manual Matcha Frother", 1,
     "UNIT", "ACTIVE", "OK", None, None, None, "FIXED", None],
    ["BOM-REPACK-MAT-KIT-L04", "BOM-REPACK-MAT-KIT", 4, "REPACK", "MATCHA", "REGULAR", "KIT",
     "RAW_NAME", "ACC-CUP-MAT-600ML", "Matcha Cup 600ML", 1,
     "UNIT", "ACTIVE", "OK", None, None, None, "FIXED", None],
    ["BOM-REPACK-MAT-KIT-L05", "BOM-REPACK-MAT-KIT", 5, "REPACK", "MATCHA", "REGULAR", "KIT",
     "RAW_NAME", "ACC-CUP-MEASURE-MAT", "Matcha Measuring Cup", 1,
     "UNIT", "ACTIVE", "OK", None, None, None, "FIXED", None],
    ["BOM-REPACK-MAT-KIT-L06", "BOM-REPACK-MAT-KIT", 6, "REPACK", "MATCHA", "REGULAR", "KIT",
     "RAW_NAME", "PKG-CARTON-MAT-KIT", "GT-branded black carton for Matcha Kit", 1,
     "UNIT", "ACTIVE", "OK", None, None, None, "FIXED", None],
]
for r_offset, row_vals in enumerate(bom_lines):
    for c_idx, val in enumerate(row_vals, start=1):
        ws.cell(row=start_row + r_offset, column=c_idx, value=val)
    print(f"  + row {start_row + r_offset}: {row_vals[0]} → {row_vals[8]} x {row_vals[10]}")

# ---------------------------------------------------------------------------
# 6. SUPPLIER_ITEM_MAP: append 2 mappings
# ---------------------------------------------------------------------------
ws = wb["SUPPLIER_ITEM_MAP"]
last_data_0 = find_last_contiguous_data(ws, 5)
start_row = last_data_0 + 2
print(f"SUPPLIER_ITEM_MAP: appending starting row {start_row}")

# Columns:
# COMPONENT_ID, COMPONENT_NAME, SUPPLIER_ID, SUPPLIER_NAME, RELATIONSHIP, IS_PRIMARY,
# ORDER_UOM, INVENTORY_UOM, UOM_CONVERSION_TO_INVENTORY, LEAD_TIME_DAYS, MOQ,
# PAYMENT_TERMS, LAST_KNOWN_PRICE, PRICE_CURRENCY, LAST_PRICE_UOM, APPROVAL_STATUS,
# SOURCE_BASIS, NOTES
supplier_maps = [
    ["PKG-BOTTLE-MAT-KIT-500ML", "Bottle 500ML for Matcha Kit", "SUP-002", "Arizot 2100",
     "PRIMARY", "YES", "UNIT", "UNIT", 1, None, None,
     None, None, "ILS", "UNIT", "PENDING",
     "User confirmed 2026-05-07", None],
    ["ACC-CUP-MAT-600ML", "Matcha Cup 600ML", "SUP-047", "Si Madeh",
     "PRIMARY", "YES", "UNIT", "UNIT", 1, None, None,
     None, None, "ILS", "UNIT", "PENDING",
     "User confirmed 2026-05-07", None],
]
for r_offset, row_vals in enumerate(supplier_maps):
    for c_idx, val in enumerate(row_vals, start=1):
        ws.cell(row=start_row + r_offset, column=c_idx, value=val)
    print(f"  + row {start_row + r_offset}: {row_vals[0]} → {row_vals[3]}")

# ---------------------------------------------------------------------------
wb.save(XLSX)
print()
print("Saved to:", XLSX)
