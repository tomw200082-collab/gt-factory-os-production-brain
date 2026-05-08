#!/usr/bin/env python3
"""Full analysis of GT_May_2026_batch_forecast_RTL (1).xlsx.

Produces a JSON report that maps every Excel row to (a) its weekly quantities,
(b) the matching item_id(s) via items.sku / items.legacy_sku / items.item_id
lookup against the live DB, and (c) the resulting forecast_lines that would
be inserted.

Outputs to scripts/_analyze_may_forecast.json so the Node-side ingest script
can read a clean structure rather than re-parsing Excel.
"""
import json
import sys
import os
import re
import urllib.parse
from pathlib import Path
from datetime import date

import openpyxl
import psycopg

ROOT = Path(r"c:/Users/tomw2/GTeveryday Dropbox/Data Center/Tom/AI Agents & Projects/Code Agents/PRODUCTION")
XLSX = ROOT / "GT_May_2026_batch_forecast_RTL (1).xlsx"
OUT = ROOT / "scripts" / "_analyze_may_forecast.json"
OUT_TXT = ROOT / "scripts" / "_analyze_may_forecast.txt"

# Read DB URL from gt-factory-os/.env
ENV = Path(r"c:/Users/tomw2/Projects/gt-factory-os/.env").read_text(encoding="utf-8")
DB_URL = ""
for line in ENV.splitlines():
    if line.startswith("DATABASE_URL_POOLED="):
        DB_URL = line[len("DATABASE_URL_POOLED="):].strip()
        break
assert DB_URL, "DATABASE_URL_POOLED not found"

# May 2026 week buckets (Sun-start, ISO Mon-Sun? File says 1-3/5, 4-10/5 etc.)
# 1-3/5 = Friday May 1, Saturday May 2, Sunday May 3 (3 days, partial week)
# 4-10/5 = Mon May 4 – Sun May 10
# 11-17/5 = Mon May 11 – Sun May 17
# 18-24/5 = Mon May 18 – Sun May 24
# 25-31/5 = Mon May 25 – Sun May 31
WEEK_BUCKETS = [
    (date(2026, 5, 1),  date(2026, 5, 3),  "Week 1 (May 1-3)"),
    (date(2026, 5, 4),  date(2026, 5, 10), "Week 2 (May 4-10)"),
    (date(2026, 5, 11), date(2026, 5, 17), "Week 3 (May 11-17)"),
    (date(2026, 5, 18), date(2026, 5, 24), "Week 4 (May 18-24)"),
    (date(2026, 5, 25), date(2026, 5, 31), "Week 5 (May 25-31)"),
]
WEEK_COLS = [13, 14, 15, 16, 17]  # 1-indexed Excel column numbers

print(f"Loading Excel: {XLSX}", file=sys.stderr)
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb[wb.sheetnames[0]]
print(f"Sheet: '{wb.sheetnames[0]}', rows={ws.max_row}, cols={ws.max_column}", file=sys.stderr)

# Pull everything from row 6 onwards
rows = []
for r in range(6, ws.max_row + 1):
    family = ws.cell(row=r, column=1).value
    name   = ws.cell(row=r, column=2).value
    flavor = ws.cell(row=r, column=3).value
    qty_total      = ws.cell(row=r, column=11).value
    qty_sales      = ws.cell(row=r, column=12).value
    weekly = []
    for c in WEEK_COLS:
        v = ws.cell(row=r, column=c).value
        weekly.append(v if isinstance(v, (int, float)) else 0)
    channel    = ws.cell(row=r, column=18).value
    seasonality= ws.cell(row=r, column=19).value
    confidence = ws.cell(row=r, column=20).value
    notes      = ws.cell(row=r, column=21).value
    skus_raw   = ws.cell(row=r, column=22).value

    if not family and not name and not skus_raw:
        continue  # blank row

    # Extract sku list (comma-separated, sometimes with spaces)
    skus = []
    if skus_raw:
        for s in re.split(r"[,;]", str(skus_raw)):
            s = s.strip().strip("'\"").strip()
            if s:
                skus.append(s)

    rows.append({
        "excel_row": r,
        "family": str(family).strip("'\"") if family else None,
        "name":   str(name).strip("'\"")   if name   else None,
        "flavor": str(flavor).strip("'\"") if flavor else None,
        "qty_total_units": float(qty_total) if isinstance(qty_total, (int,float)) else None,
        "qty_sales_units": float(qty_sales) if isinstance(qty_sales, (int,float)) else None,
        "weekly_qty":  [float(v) for v in weekly],
        "weekly_sum":  float(sum(weekly)),
        "channel":     str(channel).strip("'\"")     if channel else None,
        "seasonality": str(seasonality).strip("'\"") if seasonality else None,
        "confidence":  str(confidence).strip("'\"")  if confidence else None,
        "notes":       str(notes).strip("'\"")       if notes else None,
        "skus_raw":    str(skus_raw).strip("'\"")    if skus_raw else None,
        "skus":        skus,
    })
print(f"Parsed {len(rows)} forecast rows", file=sys.stderr)

# Connect to DB and pull items master
print("Connecting to DB...", file=sys.stderr)
conn = psycopg.connect(DB_URL, sslmode="require")
cur = conn.cursor()
cur.execute("""
  SELECT item_id, item_name, item_type, supply_method, status, sku, legacy_sku, barcode
  FROM private_core.items
  ORDER BY item_id
""")
items = []
for row in cur.fetchall():
    items.append({
        "item_id": row[0],
        "item_name": row[1],
        "item_type": row[2],
        "supply_method": row[3],
        "status": row[4],
        "sku": row[5],
        "legacy_sku": row[6],
        "barcode": row[7],
    })
print(f"Loaded {len(items)} items from DB", file=sys.stderr)

# Build lookup tables
by_item_id   = {it["item_id"].upper(): it for it in items}
by_sku       = {it["sku"].upper(): it for it in items if it.get("sku")}
by_legacy    = {it["legacy_sku"].upper(): it for it in items if it.get("legacy_sku")}

def resolve_sku(sku):
    """Try to map an Excel SKU string to a real item_id."""
    if not sku: return None, None
    key = sku.upper()
    if key in by_item_id:    return by_item_id[key]["item_id"], "item_id"
    if key in by_sku:        return by_sku[key]["item_id"], "sku"
    if key in by_legacy:     return by_legacy[key]["item_id"], "legacy_sku"
    # Try fuzzy (strip dashes/spaces)
    norm_key = re.sub(r"[\s\-]", "", key)
    for it in items:
        for col in ("item_id", "sku", "legacy_sku"):
            v = it.get(col)
            if v and re.sub(r"[\s\-]", "", v.upper()) == norm_key:
                return it["item_id"], f"{col}_normalized"
    return None, None

# Resolve every row's SKUs
report = {
    "source_file": str(XLSX.name),
    "sheet":       wb.sheetnames[0],
    "total_rows":  len(rows),
    "week_buckets": [{"start": b[0].isoformat(), "end": b[1].isoformat(), "label": b[2]} for b in WEEK_BUCKETS],
    "items_in_db": len(items),
    "rows": [],
}

unmatched_skus = set()
matched_count = 0
total_lines_to_insert = 0

for row in rows:
    resolved = []
    for sku in row["skus"]:
        item_id, match_type = resolve_sku(sku)
        resolved.append({"sku": sku, "item_id": item_id, "match_type": match_type})
        if item_id is None:
            unmatched_skus.add(sku)
        else:
            matched_count += 1

    # Forecast line generation strategy: split row's weekly_qty equally across
    # the resolved SKUs (since the Excel quantity is for the BRAND/FAMILY ROW,
    # not per-SKU). If the row has 2 SKUs (1L + 0.5L variants), each gets HALF
    # the weekly qty. Per-pack-size split using cols 4-9 is more accurate but
    # requires per-SKU logic — flagging as Plan B.
    matched_ids = [r["item_id"] for r in resolved if r["item_id"]]
    n_matched = len(matched_ids)
    lines = []
    if n_matched > 0:
        per_sku_share = 1.0 / n_matched
        for item_id in matched_ids:
            for (start, end, label), qty in zip(WEEK_BUCKETS, row["weekly_qty"]):
                qty_for_sku = qty * per_sku_share
                if qty_for_sku > 0:
                    lines.append({
                        "item_id": item_id,
                        "bucket_start": start.isoformat(),
                        "bucket_end":   end.isoformat(),
                        "qty":          qty_for_sku,
                    })
    total_lines_to_insert += len(lines)

    report["rows"].append({
        "excel_row":  row["excel_row"],
        "family":     row["family"],
        "name":       row["name"],
        "flavor":     row["flavor"],
        "weekly_qty": row["weekly_qty"],
        "weekly_sum": row["weekly_sum"],
        "qty_total_units": row["qty_total_units"],
        "qty_sales_units": row["qty_sales_units"],
        "skus_raw":  row["skus_raw"],
        "resolved":  resolved,
        "channel":   row["channel"],
        "notes":     row["notes"],
        "lines_to_insert": lines,
    })

report["unmatched_skus"]      = sorted(unmatched_skus)
report["unmatched_sku_count"] = len(unmatched_skus)
report["matched_sku_count"]   = matched_count
report["total_lines_to_insert"] = total_lines_to_insert

# Write JSON
OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote JSON: {OUT}", file=sys.stderr)

# Write human-readable report (Hebrew-safe — UTF-8)
with open(OUT_TXT, "w", encoding="utf-8") as f:
    f.write(f"=== GT May 2026 Forecast — Analysis Report ===\n\n")
    f.write(f"Source: {XLSX.name}\n")
    f.write(f"Sheet:  {wb.sheetnames[0]}\n")
    f.write(f"Excel rows parsed: {len(rows)}\n")
    f.write(f"Items in live DB:  {len(items)}\n")
    f.write(f"Distinct SKU strings in Excel: {sum(len(r['skus']) for r in rows)}\n")
    f.write(f"  → matched to item_id:  {matched_count}\n")
    f.write(f"  → unmatched:           {len(unmatched_skus)}\n")
    f.write(f"\nForecast lines to insert: {total_lines_to_insert}\n\n")

    f.write("Week buckets (May 2026):\n")
    for b in WEEK_BUCKETS:
        f.write(f"  {b[2]:24s}  {b[0]} → {b[1]}\n")
    f.write("\n")

    f.write("=" * 100 + "\n")
    f.write("PER-ROW MAPPING\n")
    f.write("=" * 100 + "\n")
    for row_data in report["rows"]:
        f.write(f"\nExcel row {row_data['excel_row']}: {row_data['family']} | {row_data['name']} | {row_data['flavor']}\n")
        f.write(f"  Weekly: {row_data['weekly_qty']}  (sum={row_data['weekly_sum']})\n")
        f.write(f"  Qty total/sales: {row_data['qty_total_units']}/{row_data['qty_sales_units']}\n")
        f.write(f"  Excel SKUs: {row_data['skus_raw']}\n")
        for r in row_data["resolved"]:
            mark = "✓" if r["item_id"] else "✗"
            f.write(f"    {mark}  {r['sku']:30s} → {r['item_id'] or '— UNMATCHED —':30s}  ({r['match_type'] or 'none'})\n")
        f.write(f"  Lines to insert from this row: {len(row_data['lines_to_insert'])}\n")
        if row_data['notes']:
            f.write(f"  Notes: {row_data['notes']}\n")

    f.write("\n" + "=" * 100 + "\n")
    f.write(f"UNMATCHED SKUs ({len(unmatched_skus)})\n")
    f.write("=" * 100 + "\n")
    for s in sorted(unmatched_skus):
        f.write(f"  {s}\n")

    f.write("\n" + "=" * 100 + "\n")
    f.write("ITEMS MASTER (for reference — what's in DB)\n")
    f.write("=" * 100 + "\n")
    for it in items:
        f.write(f"  {it['item_id']:35s}  sku={it['sku']!r:30s}  legacy_sku={it['legacy_sku']!r:30s}  name={it['item_name']}\n")

print(f"Wrote text: {OUT_TXT}", file=sys.stderr)
print(f"\nSummary:", file=sys.stderr)
print(f"  Excel rows:       {len(rows)}", file=sys.stderr)
print(f"  Total Excel SKUs: {sum(len(r['skus']) for r in rows)}", file=sys.stderr)
print(f"  Matched:          {matched_count}", file=sys.stderr)
print(f"  Unmatched:        {len(unmatched_skus)}", file=sys.stderr)
print(f"  Lines to insert:  {total_lines_to_insert}", file=sys.stderr)

cur.close()
conn.close()
