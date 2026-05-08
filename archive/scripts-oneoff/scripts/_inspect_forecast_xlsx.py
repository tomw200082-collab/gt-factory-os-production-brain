#!/usr/bin/env python3
"""Inspect GT_May_2026_batch_forecast_RTL (1).xlsx — surface every sheet,
column, and row so we can map cleanly into forecast_versions / forecast_lines."""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed; trying pandas+openpyxl alt...")
    raise

ROOT = Path(r"c:/Users/tomw2/GTeveryday Dropbox/Data Center/Tom/AI Agents & Projects/Code Agents/PRODUCTION")
PATH = ROOT / "GT_May_2026_batch_forecast_RTL (1).xlsx"

print(f"Reading: {PATH}")
print(f"Exists: {PATH.exists()}, size: {PATH.stat().st_size if PATH.exists() else 'N/A'}")
print()

wb = openpyxl.load_workbook(PATH, data_only=True)
print(f"Sheet names ({len(wb.sheetnames)}):")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  - '{s}' — {ws.max_row} rows × {ws.max_column} cols")

print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("=" * 80)
    print(f"SHEET: {sheet_name}")
    print(f"  rows={ws.max_row}, cols={ws.max_column}")
    print()
    # Print headers (first row)
    print("  Headers (row 1):")
    headers = []
    for c in range(1, min(ws.max_column, 30) + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v) if v is not None else "")
        print(f"    col {c:2d}: {repr(v)}")
    print()
    # Print first 5 data rows
    print("  First 5 data rows (row 2-6):")
    for r in range(2, min(7, ws.max_row + 1)):
        row_data = []
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(row=r, column=c).value
            row_data.append(repr(v) if v is not None else "·")
        print(f"    row {r:3d}: {row_data}")
    print()
    # Print last 3 data rows
    if ws.max_row > 7:
        print("  Last 3 data rows:")
        for r in range(max(7, ws.max_row - 2), ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column, 30) + 1):
                v = ws.cell(row=r, column=c).value
                row_data.append(repr(v) if v is not None else "·")
            print(f"    row {r:3d}: {row_data}")
    print()
