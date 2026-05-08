#!/usr/bin/env python3
"""V3 — adds Tom-approved manual SKU overrides on top of V2.

Manual overrides (Tom 2026-05-02):
  Group 1: GTCC-NM-SAN-3.85L  -> FG-NM-3850ML  (typo: NM vs NON, same product)
  Group 5: GTEL-COC-ARA-0.5L  -> FG-ARK-PAS-500ML  (Arak Passion Fruit, same product)
All other unmatched SKUs are intentionally skipped per Tom directive.
"""
import json, sys, re
from pathlib import Path
from datetime import date

import openpyxl
import psycopg

ROOT = Path(r"c:/Users/tomw2/GTeveryday Dropbox/Data Center/Tom/AI Agents & Projects/Code Agents/PRODUCTION")
XLSX = ROOT / "GT_May_2026_batch_forecast_RTL (1).xlsx"
OUT  = ROOT / "scripts" / "_analyze_may_forecast_v3.json"
OUTT = ROOT / "scripts" / "_analyze_may_forecast_v3.txt"

ENV = Path(r"c:/Users/tomw2/Projects/gt-factory-os/.env").read_text(encoding="utf-8")
DB_URL = next(l[len("DATABASE_URL_POOLED="):].strip()
              for l in ENV.splitlines() if l.startswith("DATABASE_URL_POOLED="))

# Tom-approved manual overrides
MANUAL_OVERRIDES = {
    "GTCC-NM-SAN-3.85L": "FG-NM-3850ML",
    "GTEL-COC-ARA-0.5L": "FG-ARK-PAS-500ML",
}

PACK_COLS = {
    4:  ("1L",            ["-1L",   "-1"]),
    5:  ("0.5L (500ML)",  ["-0.5L", "-500ML"]),
    6:  ("0.75L (750ML)", ["-0.75L","-750ML"]),
    7:  ("0.27/0.3L (300ML)", ["-0.3L", "-0.27L", "-300ML"]),
    8:  ("0.15/0.2L",     ["-0.15L", "-0.2L", "-150ML", "-200ML"]),
    9:  ("3.85L (3850ML)", ["-3.85L", "-3850ML"]),
}
WEEK_COLS = [13, 14, 15, 16, 17]
ISO_WEEK_STARTS = [date(2026, 5, 4), date(2026, 5, 4),
                   date(2026, 5, 11), date(2026, 5, 18), date(2026, 5, 25)]
MONTH_BUCKET = date(2026, 5, 1)

print("Loading Excel...", file=sys.stderr)
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb[wb.sheetnames[0]]

print("Connecting DB...", file=sys.stderr)
conn = psycopg.connect(DB_URL, sslmode="require")
cur = conn.cursor()
cur.execute("""SELECT item_id, item_name, item_type, supply_method, status, sku, legacy_sku
               FROM private_core.items""")
items = [{"item_id":r[0],"item_name":r[1],"item_type":r[2],"supply_method":r[3],
          "status":r[4],"sku":r[5],"legacy_sku":r[6]} for r in cur.fetchall()]

by_item_id = {it["item_id"].upper(): it for it in items}
by_sku     = {it["sku"].upper(): it for it in items if it.get("sku")}
by_legacy  = {it["legacy_sku"].upper(): it for it in items if it.get("legacy_sku")}

def resolve_sku(sku_raw):
    """Map an Excel SKU to a real item_id. Returns (item_id, match_type, matched_sku)."""
    if not sku_raw: return None, None, None
    # MANUAL OVERRIDE first
    if sku_raw in MANUAL_OVERRIDES:
        target = MANUAL_OVERRIDES[sku_raw]
        if target.upper() in by_item_id:
            return target, "manual_override", sku_raw
    candidates = [sku_raw]
    if not re.search(r"-(\d+(\.\d+)?)(ML|L|G|KG)$", sku_raw, re.IGNORECASE) and sku_raw.startswith("GT-ODK"):
        candidates += [sku_raw + "-1L", sku_raw + "-1", sku_raw + "L"]
    for cand in candidates:
        key = cand.upper()
        if key in by_item_id: return by_item_id[key]["item_id"], "item_id", cand
        if key in by_sku:     return by_sku[key]["item_id"], "sku", cand
        if key in by_legacy:  return by_legacy[key]["item_id"], "legacy_sku", cand
        norm = re.sub(r"[\s\-]", "", key)
        for it in items:
            for col in ("item_id","sku","legacy_sku"):
                v = it.get(col)
                if v and re.sub(r"[\s\-]", "", v.upper()) == norm:
                    return it["item_id"], f"{col}_normalized", cand
    return None, None, None

def detect_pack_col(sku):
    s = sku.upper()
    for col_idx, (label, suffixes) in PACK_COLS.items():
        for sfx in suffixes:
            if s.endswith(sfx.upper()):
                return col_idx, label
    return None, None

# Parse rows
rows = []
for r in range(6, ws.max_row + 1):
    family = ws.cell(row=r, column=1).value
    name   = ws.cell(row=r, column=2).value
    flavor = ws.cell(row=r, column=3).value
    pack_vals = {col: (ws.cell(row=r, column=col).value or 0) for col in PACK_COLS}
    qty_total = ws.cell(row=r, column=11).value or 0
    weekly = [ws.cell(row=r, column=c).value or 0 for c in WEEK_COLS]
    notes  = ws.cell(row=r, column=21).value
    skus_raw = ws.cell(row=r, column=22).value
    if not (family or name or skus_raw): continue
    skus = []
    if skus_raw:
        for s in re.split(r"[,;]", str(skus_raw)):
            s = s.strip().strip("'\"").strip()
            if s: skus.append(s)
    rows.append({
        "excel_row": r,
        "family": str(family).strip("'\"") if family else None,
        "name":   str(name).strip("'\"") if name else None,
        "flavor": str(flavor).strip("'\"") if flavor else None,
        "pack_vals": {str(k): float(v) for k,v in pack_vals.items()},
        "qty_total": float(qty_total),
        "weekly":   [float(v) for v in weekly],
        "weekly_sum": float(sum(weekly)),
        "notes":    str(notes).strip("'\"")   if notes else None,
        "skus_raw": str(skus_raw).strip("'\"") if skus_raw else None,
        "skus":     skus,
    })

report = {"month_bucket": MONTH_BUCKET.isoformat(),
          "iso_week_starts": [d.isoformat() for d in ISO_WEEK_STARTS],
          "items_in_db": len(items), "rows": []}

unmatched = set(); matched_distinct = set()
total_lines_monthly = 0; total_lines_weekly = 0; total_qty_assigned = 0.0

for row in rows:
    resolved = []
    for sku in row["skus"]:
        item_id, match_type, matched_sku = resolve_sku(sku)
        pack_col, pack_label = (None, None)
        if item_id:
            # Use the original SKU's pack-size suffix for share calculation
            # (manual overrides preserve the suffix from the Excel SKU).
            pack_col, pack_label = detect_pack_col(matched_sku or sku)
        resolved.append({
            "sku_excel": sku, "matched_sku": matched_sku, "item_id": item_id,
            "match_type": match_type, "pack_col": pack_col, "pack_label": pack_label,
            "pack_share_qty": float(row["pack_vals"].get(str(pack_col), 0)) if pack_col else None,
        })
        if not item_id: unmatched.add(sku)
        else: matched_distinct.add(item_id)

    monthly_lines = []; weekly_lines = []
    matched_resolved = [r for r in resolved if r["item_id"]]
    if matched_resolved:
        sum_pack_shares = sum((r["pack_share_qty"] or 0) for r in matched_resolved)
        n_matched = len(matched_resolved)
        for r in matched_resolved:
            pack_qty = r["pack_share_qty"] or 0
            if pack_qty > 0:
                qty_for_sku = pack_qty
            elif sum_pack_shares > 0:
                unassigned = max(0, row["qty_total"] - sum_pack_shares)
                count_no_pack = sum(1 for x in matched_resolved if not (x["pack_share_qty"] or 0) > 0)
                qty_for_sku = unassigned / max(count_no_pack, 1)
            else:
                qty_for_sku = row["qty_total"] / n_matched
            if qty_for_sku <= 0: continue
            monthly_lines.append({"item_id": r["item_id"],
                                  "period_bucket_key": MONTH_BUCKET.isoformat(),
                                  "forecast_quantity": round(qty_for_sku, 4)})
            if row["weekly_sum"] > 0:
                bucket_qty = {}
                for week_qty, iso_start in zip(row["weekly"], ISO_WEEK_STARTS):
                    share = week_qty / row["weekly_sum"]
                    qty_w = qty_for_sku * share
                    if qty_w > 0:
                        bucket_qty[iso_start.isoformat()] = bucket_qty.get(iso_start.isoformat(), 0) + qty_w
                for bk, q in sorted(bucket_qty.items()):
                    weekly_lines.append({"item_id": r["item_id"],
                                         "period_bucket_key": bk,
                                         "forecast_quantity": round(q, 4)})
            total_qty_assigned += qty_for_sku

    total_lines_monthly += len(monthly_lines)
    total_lines_weekly  += len(weekly_lines)
    report["rows"].append({**row, "resolved": resolved,
                           "monthly_lines": monthly_lines,
                           "weekly_lines": weekly_lines})

report["unmatched_skus"]      = sorted(unmatched)
report["unmatched_sku_count"] = len(unmatched)
report["matched_distinct_count"] = len(matched_distinct)
report["summary"] = {
    "excel_rows": len(rows),
    "total_skus_in_excel": sum(len(r["skus"]) for r in rows),
    "matched_skus": sum(len(r["skus"]) for r in rows) - len(unmatched),
    "unmatched_skus": len(unmatched),
    "manual_overrides_applied": list(MANUAL_OVERRIDES.keys()),
    "total_lines_monthly": total_lines_monthly,
    "total_lines_weekly":  total_lines_weekly,
    "total_qty_assigned":  round(total_qty_assigned, 2),
    "distinct_items_with_forecast": len(matched_distinct),
}
OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(report["summary"], indent=2), file=sys.stderr)
cur.close(); conn.close()
