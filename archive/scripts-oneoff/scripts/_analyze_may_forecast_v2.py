#!/usr/bin/env python3
"""V2 forecast analysis: per-pack-size SKU split + ODK fuzzy match + dual
cadence preview (monthly vs weekly).

Outputs JSON consumed by the Node-side ingest script (only after Tom approves).
"""
import json
import sys
import re
from pathlib import Path
from datetime import date

import openpyxl
import psycopg

ROOT = Path(r"c:/Users/tomw2/GTeveryday Dropbox/Data Center/Tom/AI Agents & Projects/Code Agents/PRODUCTION")
XLSX = ROOT / "GT_May_2026_batch_forecast_RTL (1).xlsx"
OUT  = ROOT / "scripts" / "_analyze_may_forecast_v2.json"
OUTT = ROOT / "scripts" / "_analyze_may_forecast_v2.txt"

ENV = Path(r"c:/Users/tomw2/Projects/gt-factory-os/.env").read_text(encoding="utf-8")
DB_URL = next(l[len("DATABASE_URL_POOLED="):].strip()
              for l in ENV.splitlines() if l.startswith("DATABASE_URL_POOLED="))

# Map Excel pack-size column → list of suffix strings the SKU might use
# (multiple variants because the master has both "1L" and "1L" patterns).
PACK_COLS = {
    4:  ("1L",            ["-1L",   "-1"]),
    5:  ("0.5L (500ML)",  ["-0.5L", "-500ML"]),
    6:  ("0.75L (750ML)", ["-0.75L","-750ML"]),
    7:  ("0.27/0.3L (300ML)", ["-0.3L", "-0.27L", "-300ML"]),
    8:  ("0.15/0.2L (150-200ML)", ["-0.15L", "-0.2L", "-150ML", "-200ML"]),
    9:  ("3.85L (3850ML)", ["-3.85L", "-3850ML"]),
}

WEEK_COLS = [13, 14, 15, 16, 17]
WEEK_BUCKETS = [
    (date(2026, 5, 4),  "Week 1+2 (May 1-10) — partial+full"),  # combine Week 1 (3 days) into Week 2 ISO
    (date(2026, 5, 4),  "Week 2 (May 4-10)"),
    (date(2026, 5, 11), "Week 3 (May 11-17)"),
    (date(2026, 5, 18), "Week 4 (May 18-24)"),
    (date(2026, 5, 25), "Week 5 (May 25-31)"),
]
# Use Mon-start week-start dates per ISO. Week 1 Friday is folded into Week 2.
ISO_WEEK_STARTS = [
    date(2026, 5, 4),   # Week 1 (May 1-3, partial) folded into here
    date(2026, 5, 4),   # Week 2 (4-10)
    date(2026, 5, 11),  # Week 3
    date(2026, 5, 18),  # Week 4
    date(2026, 5, 25),  # Week 5
]
MONTH_BUCKET = date(2026, 5, 1)

print(f"Loading Excel...", file=sys.stderr)
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb[wb.sheetnames[0]]

print("Connecting to DB...", file=sys.stderr)
conn = psycopg.connect(DB_URL, sslmode="require")
cur = conn.cursor()
cur.execute("""
  SELECT item_id, item_name, item_type, supply_method, status, sku, legacy_sku
  FROM private_core.items
  ORDER BY item_id
""")
items = [{"item_id":r[0],"item_name":r[1],"item_type":r[2],"supply_method":r[3],
          "status":r[4],"sku":r[5],"legacy_sku":r[6]} for r in cur.fetchall()]
print(f"Loaded {len(items)} items", file=sys.stderr)

by_item_id = {it["item_id"].upper(): it for it in items}
by_sku     = {it["sku"].upper(): it for it in items if it.get("sku")}
by_legacy  = {it["legacy_sku"].upper(): it for it in items if it.get("legacy_sku")}

def resolve_sku(sku_raw):
    """Map an Excel SKU to a real item. Returns (item_id, match_type, matched_sku)."""
    if not sku_raw: return None, None, None
    candidates = [sku_raw]
    # Smart variants: ODK-style (no pack-size suffix) → try adding common suffixes
    if not re.search(r"-(\d+(\.\d+)?)(ML|L|G|KG)$", sku_raw, re.IGNORECASE) and sku_raw.startswith("GT-ODK"):
        candidates += [sku_raw + "-1L", sku_raw + "-1", sku_raw + "L"]
    for cand in candidates:
        key = cand.upper()
        if key in by_item_id:    return by_item_id[key]["item_id"], "item_id", cand
        if key in by_sku:        return by_sku[key]["item_id"], "sku", cand
        if key in by_legacy:     return by_legacy[key]["item_id"], "legacy_sku", cand
        norm = re.sub(r"[\s\-]", "", key)
        for it in items:
            for col in ("item_id","sku","legacy_sku"):
                v = it.get(col)
                if v and re.sub(r"[\s\-]", "", v.upper()) == norm:
                    return it["item_id"], f"{col}_normalized", cand
    return None, None, None

def detect_pack_col(sku):
    """Given an SKU like GT-LUI-LOW-1L, return the Excel pack-size column index."""
    s = sku.upper()
    for col_idx, (label, suffixes) in PACK_COLS.items():
        for sfx in suffixes:
            if s.endswith(sfx.upper()):
                return col_idx, label
    return None, None

# Parse rows
rows = []
for r in range(6, ws.max_row + 1):
    family = ws.cell(row=r, column=1).value
    name   = ws.cell(row=r, column=2).value
    flavor = ws.cell(row=r, column=3).value
    pack_vals = {col: (ws.cell(row=r, column=col).value or 0) for col in PACK_COLS}
    qty_total = ws.cell(row=r, column=11).value or 0
    qty_sales = ws.cell(row=r, column=12).value or 0
    weekly = [ws.cell(row=r, column=c).value or 0 for c in WEEK_COLS]
    channel = ws.cell(row=r, column=18).value
    notes   = ws.cell(row=r, column=21).value
    skus_raw = ws.cell(row=r, column=22).value
    if not (family or name or skus_raw): continue

    skus = []
    if skus_raw:
        for s in re.split(r"[,;]", str(skus_raw)):
            s = s.strip().strip("'\"").strip()
            if s: skus.append(s)

    rows.append({
        "excel_row": r,
        "family": str(family).strip("'\"") if family else None,
        "name":   str(name).strip("'\"") if name else None,
        "flavor": str(flavor).strip("'\"") if flavor else None,
        "pack_vals": {str(k): float(v) for k,v in pack_vals.items()},
        "qty_total": float(qty_total),
        "qty_sales": float(qty_sales) if isinstance(qty_sales,(int,float)) else None,
        "weekly":   [float(v) for v in weekly],
        "weekly_sum": float(sum(weekly)),
        "channel":  str(channel).strip("'\"") if channel else None,
        "notes":    str(notes).strip("'\"")   if notes else None,
        "skus_raw": str(skus_raw).strip("'\"") if skus_raw else None,
        "skus":     skus,
    })

# Per-row: resolve each SKU, determine its pack-size share, generate forecast lines
report = {
    "source_file":  XLSX.name,
    "sheet":        wb.sheetnames[0],
    "items_in_db":  len(items),
    "month_bucket": MONTH_BUCKET.isoformat(),
    "iso_week_starts": [d.isoformat() for d in ISO_WEEK_STARTS],
    "rows": [],
    "unmatched_skus": [],
    "summary": {},
}

unmatched = set()
matched_distinct = set()
total_lines_monthly = 0
total_lines_weekly = 0
total_qty_assigned = 0.0

for row in rows:
    resolved = []
    for sku in row["skus"]:
        item_id, match_type, matched_sku = resolve_sku(sku)
        pack_col, pack_label = (None, None)
        if item_id:
            pack_col, pack_label = detect_pack_col(matched_sku or sku)
        resolved.append({
            "sku_excel": sku,
            "matched_sku": matched_sku,
            "item_id": item_id,
            "match_type": match_type,
            "pack_col": pack_col,
            "pack_label": pack_label,
            "pack_share_qty": float(row["pack_vals"].get(str(pack_col), 0)) if pack_col else None,
        })
        if not item_id: unmatched.add(sku)
        else: matched_distinct.add(item_id)

    # Generate monthly lines (one per matched SKU, qty = pack-size col value)
    monthly_lines = []
    weekly_lines  = []
    matched_resolved = [r for r in resolved if r["item_id"]]

    if matched_resolved:
        # Sum of pack-share over matched SKUs (used for proportional split when
        # pack-size detection failed for one SKU but worked for others)
        sum_pack_shares = sum((r["pack_share_qty"] or 0) for r in matched_resolved)
        n_matched = len(matched_resolved)

        for r in matched_resolved:
            pack_qty = r["pack_share_qty"] or 0
            if pack_qty > 0:
                # Use pack-size column value as the SKU's monthly total
                qty_for_sku = pack_qty
            elif sum_pack_shares > 0:
                # If this SKU had no pack-size detected but others did, give it
                # the residual (qty_total - sum_pack_shares) / count_unassigned
                unassigned = max(0, row["qty_total"] - sum_pack_shares)
                # Simplification: give residual to ALL unmatched-pack SKUs equally
                count_no_pack = sum(1 for x in matched_resolved if not (x["pack_share_qty"] or 0) > 0)
                qty_for_sku = unassigned / max(count_no_pack, 1)
            else:
                # No pack-size info at all → equal split
                qty_for_sku = row["qty_total"] / n_matched

            if qty_for_sku <= 0:
                continue

            # Monthly: one bucket
            monthly_lines.append({
                "item_id": r["item_id"],
                "period_bucket_key": MONTH_BUCKET.isoformat(),
                "forecast_quantity": round(qty_for_sku, 4),
            })

            # Weekly: split by weekly_qty / weekly_sum × qty_for_sku
            if row["weekly_sum"] > 0:
                # Aggregate weekly buckets by ISO week start (May 1-3 partial folds into May 4)
                bucket_qty = {}
                for week_qty, iso_start in zip(row["weekly"], ISO_WEEK_STARTS):
                    share = week_qty / row["weekly_sum"]
                    qty_w = qty_for_sku * share
                    if qty_w > 0:
                        bucket_qty[iso_start.isoformat()] = bucket_qty.get(iso_start.isoformat(), 0) + qty_w
                for bk, q in sorted(bucket_qty.items()):
                    weekly_lines.append({
                        "item_id": r["item_id"],
                        "period_bucket_key": bk,
                        "forecast_quantity": round(q, 4),
                    })
            total_qty_assigned += qty_for_sku

    total_lines_monthly += len(monthly_lines)
    total_lines_weekly  += len(weekly_lines)

    report["rows"].append({
        "excel_row": row["excel_row"],
        "family":    row["family"],
        "name":      row["name"],
        "flavor":    row["flavor"],
        "pack_vals": row["pack_vals"],
        "qty_total": row["qty_total"],
        "qty_sales": row["qty_sales"],
        "weekly":    row["weekly"],
        "weekly_sum": row["weekly_sum"],
        "skus_raw":  row["skus_raw"],
        "resolved":  resolved,
        "monthly_lines": monthly_lines,
        "weekly_lines":  weekly_lines,
        "channel": row["channel"],
        "notes":   row["notes"],
    })

report["unmatched_skus"]      = sorted(unmatched)
report["unmatched_sku_count"] = len(unmatched)
report["matched_distinct_items"] = sorted(matched_distinct)
report["matched_distinct_count"] = len(matched_distinct)
report["summary"] = {
    "excel_rows": len(rows),
    "total_skus_in_excel": sum(len(r["skus"]) for r in rows),
    "matched_skus": sum(len(r["skus"]) for r in rows) - len(unmatched),
    "unmatched_skus": len(unmatched),
    "total_lines_monthly": total_lines_monthly,
    "total_lines_weekly":  total_lines_weekly,
    "total_qty_assigned":  round(total_qty_assigned, 2),
}

OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

# Human report
with open(OUTT, "w", encoding="utf-8") as f:
    s = report["summary"]
    f.write(f"=== GT May 2026 Forecast — V2 Analysis ===\n\n")
    f.write(f"Source: {XLSX.name}\n")
    f.write(f"Items in DB: {report['items_in_db']}\n\n")
    f.write(f"Excel rows: {s['excel_rows']}\n")
    f.write(f"Total SKUs in Excel: {s['total_skus_in_excel']}\n")
    f.write(f"  Matched: {s['matched_skus']}\n")
    f.write(f"  Unmatched: {s['unmatched_skus']}\n\n")
    f.write(f"Distinct items that will receive forecast: {report['matched_distinct_count']}\n\n")
    f.write(f"Total qty assigned (sum of all forecast_quantity): {s['total_qty_assigned']}\n\n")
    f.write(f"Forecast lines per cadence option:\n")
    f.write(f"  MONTHLY (1 bucket = 2026-05-01): {s['total_lines_monthly']} lines\n")
    f.write(f"  WEEKLY (4 buckets May 4/11/18/25): {s['total_lines_weekly']} lines\n\n")
    f.write("=" * 100 + "\nPER-ROW DETAIL\n" + "=" * 100 + "\n")
    for row in report["rows"]:
        f.write(f"\nRow {row['excel_row']}: {row['family']} | {row['name']} | {row['flavor']}\n")
        f.write(f"  Pack-size totals: 1L={row['pack_vals'].get('4',0)}, 0.5L={row['pack_vals'].get('5',0)}, "
                f"0.75L={row['pack_vals'].get('6',0)}, 0.3L={row['pack_vals'].get('7',0)}, "
                f"0.15-0.2L={row['pack_vals'].get('8',0)}, 3.85L={row['pack_vals'].get('9',0)}\n")
        f.write(f"  Total units: {row['qty_total']}, Sales units: {row['qty_sales']}\n")
        f.write(f"  Weekly: {row['weekly']} (sum={row['weekly_sum']})\n")
        f.write(f"  Excel SKUs: {row['skus_raw']}\n")
        for r in row["resolved"]:
            mark = "OK" if r["item_id"] else "X "
            pack = f"  pack-col={r['pack_col']} ({r['pack_label']}) qty={r['pack_share_qty']}" if r["pack_col"] else ""
            f.write(f"    {mark}  {r['sku_excel']:30s} -> {r['item_id'] or '— UNMATCHED —':35s}  ({r['match_type'] or 'none'}){pack}\n")
        if row["monthly_lines"]:
            f.write(f"  MONTHLY lines ({len(row['monthly_lines'])}):\n")
            for l in row["monthly_lines"]:
                f.write(f"    {l['item_id']:35s}  bucket={l['period_bucket_key']}  qty={l['forecast_quantity']}\n")
        if row["weekly_lines"]:
            f.write(f"  WEEKLY lines ({len(row['weekly_lines'])}):\n")
            for l in row["weekly_lines"]:
                f.write(f"    {l['item_id']:35s}  bucket={l['period_bucket_key']}  qty={l['forecast_quantity']}\n")
        if row["notes"]:
            f.write(f"  Notes: {row['notes']}\n")
    f.write("\n" + "=" * 100 + "\n")
    f.write(f"UNMATCHED SKUs ({len(unmatched)}):\n")
    for s2 in sorted(unmatched):
        f.write(f"  {s2}\n")

print("DONE", file=sys.stderr)
print(json.dumps(report["summary"], indent=2), file=sys.stderr)
cur.close(); conn.close()
