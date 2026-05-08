"""
Build May 2026 procurement plan workbook.

Inputs:
  - GT_May_2026_batch_forecast_RTL (1).xlsx  (forecast, 57 batches)
  - GT_Master_Data.xlsx  (Item Master, BOM_HEADER, BOM_LINES, per-SKU sheets)

Output:
  - GT_May_2026_Purchase_Plan.xlsx
    Sheet 1: סיכום רכש  (aggregated buy list across all batches)
    Sheets 2..N: one sheet per batch in scope with editable forecast inputs
                 and live formulas for packaging + raw materials.

Approach:
  - Forecast quantities per pack-size are EDITABLE input cells.
  - Total liters = SUM(units * fill).
  - Packaging needs = units_per_size * pack_BOM_qty (formula references input cell).
  - Raw materials = total_liters * per_liter_ratio (precomputed in Python from
    fully-exploded base BOMs incl. nested base mixes; ratio is a constant in
    the formula so the chain stays transparent).
  - Summary aggregates each unique component from each per-batch sheet via
    explicit cross-sheet SUMIFS, so editing any forecast input recomputes
    the entire workbook.
"""

import os
import re
from collections import OrderedDict, defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string

# ────────────────────────────────────────────────────────────────────────────
# Paths
# ────────────────────────────────────────────────────────────────────────────
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FORECAST_FILE = os.path.join(ROOT, "GT_May_2026_batch_forecast_RTL (1).xlsx")
MASTER_FILE = os.path.join(ROOT, "GT_Master_Data.xlsx")
OUT_FILE = os.path.join(ROOT, "GT_May_2026_Purchase_Plan.xlsx")
STOCK_TEMPLATE_FILE = os.path.join(ROOT, "GT_May_2026_Stock_Fill_Template.xlsx")


# ────────────────────────────────────────────────────────────────────────────
# Recipe overrides (Tom-supplied corrections to the master recipes)
# ────────────────────────────────────────────────────────────────────────────
# NAMASTEA (new) — from "Namastea (new)" in cost-of-production-Aug-2025 file.
# Replaces the master's BOM-BASE-NAM-REG which used the pre-mixed 25 kg
# Chai Masala. New recipe explodes the masala into individual spices and
# bumps puer/cinnamon/etc. Output volume is 480 L per batch (master had 492 L).
NAMASTEA_NEW = {
    "BOM-BASE-NAM-REG": {
        "total_qty": 480,  # L
        "lines": [
            {"comp_id": "RAW-WATER", "name": "Water", "qty": 420, "uom": "L",
             "supplier": "Internal Plant Supply"},
            {"comp_id": "RAW-CLOVE", "name": "Whole Clove", "qty": 2.9, "uom": "KG",
             "supplier": "TAVLINEI BAR"},
            {"comp_id": "RAW-PAPER", "name": "Black Pepper", "qty": 2.4, "uom": "KG",
             "supplier": "Tavlinei Bar"},
            {"comp_id": "RAW-EL", "name": "Green Cardamom Pods", "qty": 2.21, "uom": "KG",
             "supplier": "Tavlinei Bar"},
            {"comp_id": "RAW-GINGER-CRUSHED", "name": "Crushed Ginger", "qty": 0.95, "uom": "KG",
             "supplier": "Tavlinei Bar"},
            {"comp_id": "RAW-CINNAMON", "name": "Cinnamon (Crushed)", "qty": 13.54, "uom": "KG",
             "supplier": "Tavlinei Bar"},
            {"comp_id": "RAW-BLACK-TEA", "name": "Black Tea", "qty": 6, "uom": "KG",
             "supplier": "AMITEA"},
            {"comp_id": "RAW-PUER", "name": "Puerh Tea", "qty": 4, "uom": "KG",
             "supplier": "AMITEA"},
            {"comp_id": "RAW-SUGAR", "name": "Sugar", "qty": 150, "uom": "KG",
             "supplier": "GT Everyday"},
            {"comp_id": "RAW-PRESERVATIVE", "name": "Preservative (Bulk)", "qty": 0.5, "uom": "L",
             "supplier": "Ziv Chemically"},
            {"comp_id": "RAW-STABILISER", "name": "Stabilizer", "qty": 0.1, "uom": "KG",
             "supplier": "Ziv Chemically"},
            {"comp_id": "PKG-FILTER-100MICRON", "name": "Single Length Filter 100 micron",
             "qty": 1, "uom": "UNIT", "supplier": 'א.ד טכנולוגיות סינון בע"מ'},
            {"comp_id": "PKG-FILTER-200MICRON", "name": "Double Length Filter 200 micron",
             "qty": 1, "uom": "UNIT", "supplier": 'א.ד טכנולוגיות סינון בע"מ'},
        ],
    },
}


def apply_recipe_overrides(master):
    """Replace named base recipes with hand-coded corrections."""
    for bom_id, rec in NAMASTEA_NEW.items():
        master["base_recipes"][bom_id] = rec["lines"]
        master["base_recipe_total"][bom_id] = rec["total_qty"]
        master["base_outputs"][bom_id] = rec["total_qty"]


# ────────────────────────────────────────────────────────────────────────────
# Master data loading
# ────────────────────────────────────────────────────────────────────────────
def load_master():
    wb = load_workbook(MASTER_FILE, data_only=True)

    # ITEM_MASTER
    items = {}
    items_by_legacy = {}
    ws = wb["ITEM_MASTER"]
    hdr = None
    for r in ws.iter_rows(values_only=True):
        if r[0] == "ITEM_ID":
            hdr = list(r)
            continue
        if hdr and r[0]:
            d = dict(zip(hdr, r))
            items[d["ITEM_ID"]] = d
            if d.get("LEGACY_SKU"):
                items_by_legacy[d["LEGACY_SKU"]] = d

    # BOM_HEADER → output qty per base
    base_outputs = {}
    ws = wb["BOM_HEADER"]
    hdr = None
    for r in ws.iter_rows(values_only=True):
        if r and r[0] == "BOM_ID":
            hdr = list(r)
            continue
        if hdr and r and r[0]:
            d = dict(zip(hdr, r))
            if "BASE" in str(d.get("BOM_KIND", "")):
                base_outputs[d["BOM_ID"]] = float(d["FINAL_BOM_OUTPUT_QTY"])

    # Per-SKU sheets → Pack BOM and Base BOM lines
    per_sku = {}  # sku → {pack: [...], base: [...], pack_bom_id, base_bom_id, fill}
    base_recipes = {}  # base_bom_id → list of {comp_id, name, qty, uom, supplier}
    base_recipe_total = {}  # base_bom_id → total batch qty (e.g. 500L)

    sku_sheets = [
        s
        for s in wb.sheetnames
        if s
        not in {
            "INDEX",
            "ITEM_MASTER",
            "COMPONENT_MASTER",
            "BOM_HEADER",
            "BOM_LINES",
            "SUPPLIER_MASTER",
            "SUPPLIER_ITEM_MAP",
            "FG_STOCK",
            "RM_STOCK",
            "OPEN_SUPPLY",
            "OPEN_GAPS",
            "PRICE_HISTORY",
            "CHANGE_LOG",
            "_META",
        }
    ]

    for sn in sku_sheets:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        sku = None
        pack_bom_id = None
        base_bom_id = None
        fill = None
        section = None
        pack_lines = []
        base_lines = []
        base_total_qty = None

        for r in rows:
            if not r or all(c is None for c in r):
                continue
            c0 = r[0]
            c0s = str(c0) if c0 is not None else ""
            if c0s == "SKU":
                sku = r[1]
                pack_bom_id = r[3]
                base_bom_id = r[5] if r[5] not in (None, "--") else None
                try:
                    fill = float(r[7]) if r[7] is not None else 0
                except (TypeError, ValueError):
                    fill = 0
                continue
            if "PACK BOM" in c0s:
                section = "pack"
                continue
            if "BASE BOM" in c0s:
                section = "base"
                continue
            if c0s in ("Component ID", "Raw Material ID"):
                continue
            up = c0s.upper()
            if "COST" in up or up.startswith("TOTAL") or c0s.startswith("GT Everyday"):
                continue
            if c0s.startswith("Shared base"):
                continue
            if c0s.startswith("(no base"):
                continue

            comp_id = c0
            name = r[1]
            qty = r[2]
            uom = r[3]
            supplier = r[4]

            if comp_id is None or not isinstance(comp_id, str) or not comp_id.strip():
                continue

            # Skip the first line of [PACK] which restates the base BOM at fill volume
            if section == "pack" and comp_id.startswith("BOM-BASE-"):
                continue

            try:
                qty_f = float(qty) if qty is not None else 0
            except (TypeError, ValueError):
                continue
            if qty_f == 0:
                continue

            line = {
                "comp_id": comp_id,
                "name": name or "",
                "qty": qty_f,
                "uom": (uom or "").strip(),
                "supplier": (supplier or "").strip(),
            }
            if section == "pack":
                pack_lines.append(line)
            elif section == "base":
                base_lines.append(line)

        if sku:
            per_sku[sku] = {
                "pack_bom_id": pack_bom_id,
                "base_bom_id": base_bom_id,
                "fill": fill,
                "pack": pack_lines,
                "base": base_lines,
            }
            if base_bom_id and base_bom_id not in base_recipes:
                # Use this sheet's [BASE] section as the canonical recipe
                base_recipes[base_bom_id] = base_lines
                # The base recipe is at base_outputs[base_bom_id] scale
                base_recipe_total[base_bom_id] = base_outputs.get(base_bom_id)

    return {
        "items": items,
        "items_by_legacy": items_by_legacy,
        "per_sku": per_sku,
        "base_outputs": base_outputs,
        "base_recipes": base_recipes,
        "base_recipe_total": base_recipe_total,
    }


# ────────────────────────────────────────────────────────────────────────────
# Recursive base-BOM explosion → root raws and per-batch consumables
# ────────────────────────────────────────────────────────────────────────────
def explode_base(base_bom_id, master, _depth=0):
    """
    Return dict[comp_id] → {name, uom, supplier, qty_per_liter_of_this_base}.

    Walks nested BOM-BASE-* lines. For each leaf component, sums up the
    contribution-per-liter of THIS base BOM.
    """
    if _depth > 8:
        raise RuntimeError(f"BOM nesting too deep at {base_bom_id}")
    recipe = master["base_recipes"].get(base_bom_id, [])
    total = master["base_recipe_total"].get(base_bom_id)
    if not recipe or not total:
        return {}

    out = OrderedDict()
    for line in recipe:
        comp_id = line["comp_id"]
        qty = line["qty"]  # absolute qty per `total` of this base
        per_liter = qty / total

        if isinstance(comp_id, str) and comp_id.startswith("BOM-BASE-"):
            sub = explode_base(comp_id, master, _depth + 1)
            # `qty` here is liters of sub-base consumed per `total` liters of this base
            # so per liter of THIS base we consume per_liter of sub-base
            for sub_id, sub_info in sub.items():
                key = sub_id
                contrib = per_liter * sub_info["qty_per_l"]
                if key in out:
                    out[key]["qty_per_l"] += contrib
                else:
                    out[key] = {
                        "name": sub_info["name"],
                        "uom": sub_info["uom"],
                        "supplier": sub_info["supplier"],
                        "qty_per_l": contrib,
                    }
        else:
            key = comp_id
            if key in out:
                out[key]["qty_per_l"] += per_liter
            else:
                out[key] = {
                    "name": line["name"],
                    "uom": line["uom"],
                    "supplier": line["supplier"],
                    "qty_per_l": per_liter,
                }
    return out


# ────────────────────────────────────────────────────────────────────────────
# Forecast loading & filtering
# ────────────────────────────────────────────────────────────────────────────
EXCLUDED_SKUS_OR_NAMES = {
    # cocktails to exclude (per Tom's image)
    "GT Elita Arak", "GTEL-COC-ARA-0.5L",
    "קוקטייל סנגריה ורודה 0.15 ל",
    "GT Elita White Sangria", "GTEL-SAN-WHI-0.75L",
    "קוקטייל סלסול 0.2 ל", "12002",
    "GT Elita Cosmo Lychee", "GTEL-COS-LYC-0.3L",
    "Muza Jasmine", "GTCC-MUZ-JAS-0.2L",
    "קוקטייל תה תפוחים ירושלמי (ערק אמרטו) 0.15 ל", "GTCC-JER-ARA-0.15L",
    "Muza Herbal Mule Bliss", "GTCC-MUZ-HER-0.2L",
    "Muza Negroni", "GTCC-MUZ-NEG-0.2L",
    "קוקטייל קמפרי טי דאנס (קמפרי) 0.15 ל", "GTCC-CAM-DAN-0.15L",
    "Muza Queen Violet", "GTCC-MUZ-QUE-0.2L",
    "Muza Pink Mama", "GTCC-MUZ-PNMM-1L",
    "קוקטייל טרופיקל אין גאפן (גין) 1 ל", "GTCC-TRO-JAP-1L",
    "קוקטייל ראשן, וארי ראשן (וודקה) 0.15 ל", "GTCC-VAR-RAS-0.15L",
    "GT Babka Bakery Red Sangria", "GTEL-BAB-RED-0.75L",
    "GT Babka Bakery White Sangria", "GTEL-BAB-WHI-0.75L",
    "Muza Apple Zest", "GTCC-MUZ-APPZ-1L",
    "Muza Spicy Margarita", "GTCC-MUZ-SMAR-1L",
    "Muza Tropical in Japan", "GTCC-MUZ-TROJ-1L",
    "קוקטייל גאסמין גאז (וויסקי ) 0.15 ל", "GTCC-JAS-JAZ-0.15L",
    "Muza Passion Spritz",  # both 1L and 0.2L variants share name
    "GTCC-MUZ-PSSP-1L", "GTCC-MUZ-PSC-0.2L",
    "Muza Anise Bliss", "GTCC-MUZ-ANBL-1L",
    "Muza Blueberry Breeze", "GTCC-MUZ-BLBR-1L",
    "GT Sangria", "GTCC-NM-SAN-3.85L",
    "Muza Jasmin", "GTCC-MUZ-JASM-1L",
    # all mixers
    "Muza Pink Mama Mixer", "GTMX-MUZ-PNMM-1L",
    "Muza Jasmin Mixer", "GTMX-MUZ-JASM-1L",
    "Muza Tropical Israeli Mixer", "GTMX-MUZ-TRIL-1L",
    "Muza Basil Smash Mixer", "GTMX-MUZ-BZSM-1L",
    "Muza Classic Margarita Mixer", "GTMX-MUZ-MRCL-1L",
    "Muza Herbal Mule Bliss Mixer", "GTMX-MUZ-HER-1L",
    # ODK
    "ODK Strawberry Smoothie", "ODK Mango Smoothie", "ODK Peach Smoothie",
    # Other
    "GT Arak Lemons", "GTCC-ARA-LEM-0.27L",
    # Matcha kit (deleted per Tom)
    "Complete Matcha Kit", "GT-MAT-KIT",
}


def load_forecast():
    wb = load_workbook(FORECAST_FILE, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    # Header row is index 4 (zero-based) — has 'סוג', 'באצ׳ / מוצר', ...
    hdr = rows[4]
    data_rows = []
    for r in rows[5:]:
        if not r or not r[0]:
            continue
        if r[0] == "סה״כ":
            continue
        d = dict(zip(hdr, r))
        data_rows.append(d)
    return data_rows


# Pack-size column lookup in forecast row dict
PACK_COLS = {
    "1L": "1L",
    "0.5L": "0.5L",
    "0.75L": "0.75L",
    "0.27/0.3L": "0.27/0.3L",
    "0.15/0.2L": "0.15/0.2L",
    "3.85L": "3.85L",
    "אחר": "אחר",
}


# ────────────────────────────────────────────────────────────────────────────
# Build batch list
# ────────────────────────────────────────────────────────────────────────────
def build_scope(forecast_rows, master):
    """
    Return list of batch dicts in plan order:
      {
        sheet_name, hebrew_name, family_label, base_bom_id,
        liquid: True/False,
        pack_inputs: [(label, units, fill, sku, pack_lines)],
        raws_per_l: dict (only for liquid batches),
        notes: str
      }
    """
    batches = []

    # 1) The 10 tea/herbal batches (each has 1L + 0.5L)
    tea_specs = [
        ("DETOX", "DETOX", "GT-LUI-LOW-1L", "GT-LUI-LOW-0.5L"),
        ("NAMASTEA", "NAMASTEA", "GT-MAS-CHA-1L", "GT-MAS-CHA-0.5L"),
        ("FRESH", "FRESH", "GT-HIB-LOW-1L", "GT-HIB-LOW-0.5L"),
        ("REVIVE", "REVIVE", "GT-SEN-LOW-1L", "GT-SEN-LOW-0.5L"),
        ("DETOX Sugar-Free", "DETOX SF", "GT-LUI-FRE-1L", "GT-LUI-FRE-0.5L"),
        ("ENERGY", "ENERGY", "GT-LEM-LOW-1L", "GT-LEM-LOW-0.5L"),
        ("CALM", "CALM", "GT-CHA-LOW-1L", "GT-CHA-LOW-0.5L"),
        ("CONSCIOUSNESS", "CONSCIOUSNESS", "GT-JAS-LOW-1L", "GT-JAS-LOW-0.5L"),
        ("FRESH Sugar-Free", "FRESH SF", "GT-HIB-FRE-1L", "GT-HIB-FRE-0.5L"),
        ("Desert Infusion", "DESERT INFUSION", "GT-INF-DES-1L", None),
    ]

    fc_by_batch_name = {r["באצ׳ / מוצר"]: r for r in forecast_rows if r.get("באצ׳ / מוצר")}

    for fc_name, sheet_label, lid_1l, lid_05l in tea_specs:
        fc = fc_by_batch_name.get(fc_name)
        if not fc:
            continue
        units_1l = float(fc.get("1L") or 0)
        units_05l = float(fc.get("0.5L") or 0)

        item_1l = master["items_by_legacy"].get(lid_1l)
        item_05l = master["items_by_legacy"].get(lid_05l) if lid_05l else None

        sku_1l = item_1l["ITEM_ID"] if item_1l else None
        sku_05l = item_05l["ITEM_ID"] if item_05l else None
        base_bom_id = item_1l["BASE_BOM_ID"]

        pack_inputs = []
        if item_1l:
            pack_inputs.append({
                "label": "1L",
                "fill": 1.0,
                "default_units": units_1l,
                "sku": sku_1l,
                "pack_lines": master["per_sku"][sku_1l]["pack"],
            })
        if item_05l:
            pack_inputs.append({
                "label": "0.5L",
                "fill": 0.5,
                "default_units": units_05l,
                "sku": sku_05l,
                "pack_lines": master["per_sku"][sku_05l]["pack"],
            })

        raws = explode_base(base_bom_id, master)

        batches.append({
            "sheet_name": sheet_label,
            "hebrew_name": sheet_label,
            "family_label": sheet_label,
            "base_bom_id": base_bom_id,
            "liquid": True,
            "pack_inputs": pack_inputs,
            "raws_per_l": raws,
            "notes": "",
        })

    # 2) Cocktails
    # GT Elita Red Sangria — 192 units of 0.75L
    fc = fc_by_batch_name.get("GT Elita Red Sangria")
    if fc:
        units_075 = float(fc.get("0.75L") or 0)
        item = master["items_by_legacy"].get("GTEL-SAN-RED-0.75L")
        sku = item["ITEM_ID"]
        base_bom_id = item["BASE_BOM_ID"]
        raws = explode_base(base_bom_id, master)
        batches.append({
            "sheet_name": "RED SANGRIA",
            "hebrew_name": "סנגריה אדומה (Elita)",
            "family_label": "סנגריה אדומה (Elita)",
            "base_bom_id": base_bom_id,
            "liquid": True,
            "pack_inputs": [{
                "label": "0.75L",
                "fill": 0.75,
                "default_units": units_075,
                "sku": sku,
                "pack_lines": master["per_sku"][sku]["pack"],
            }],
            "raws_per_l": raws,
            "notes": "",
        })

    # Nonomimi Sangria — 89 units 1L + 29 units 3.85L
    fc = fc_by_batch_name.get("GT Nonomimi Sangria")
    if fc:
        units_1l = float(fc.get("1L") or 0)
        units_385 = float(fc.get("3.85L") or 0)
        item_1l = master["items_by_legacy"].get("GTCC-NON-SAN-1L")
        item_385 = master["items_by_legacy"].get("GTCC-NON-SAN-3.85L")
        base_bom_id = item_1l["BASE_BOM_ID"]
        raws = explode_base(base_bom_id, master)
        pack_inputs = [
            {"label": "1L", "fill": 1.0, "default_units": units_1l,
             "sku": item_1l["ITEM_ID"],
             "pack_lines": master["per_sku"][item_1l["ITEM_ID"]]["pack"]},
            {"label": "3.85L", "fill": 3.85, "default_units": units_385,
             "sku": item_385["ITEM_ID"],
             "pack_lines": master["per_sku"][item_385["ITEM_ID"]]["pack"]},
        ]
        batches.append({
            "sheet_name": "NONOMIMI SANGRIA",
            "hebrew_name": "סנגריה נונומימי",
            "family_label": "סנגריה נונומימי",
            "base_bom_id": base_bom_id,
            "liquid": True,
            "pack_inputs": pack_inputs,
            "raws_per_l": raws,
            "notes": "",
        })

    # White Sangria — 22 units 1L (skipping 0.15L which has no SKU)
    fc = fc_by_batch_name.get("קוקטייל סנגריה לבנה 0.15 ל")
    if fc:
        units_1l = float(fc.get("1L") or 0)
        item = master["items_by_legacy"].get("GTCC-WHI-SAN-1L")
        sku = item["ITEM_ID"]
        base_bom_id = item["BASE_BOM_ID"]
        raws = explode_base(base_bom_id, master)
        batches.append({
            "sheet_name": "WHITE SANGRIA",
            "hebrew_name": "סנגריה לבנה",
            "family_label": "סנגריה לבנה",
            "base_bom_id": base_bom_id,
            "liquid": True,
            "pack_inputs": [{
                "label": "1L",
                "fill": 1.0,
                "default_units": units_1l,
                "sku": sku,
                "pack_lines": master["per_sku"][sku]["pack"],
            }],
            "raws_per_l": raws,
            "notes": "0.15L: 112 יחידות בתחזית — אין SKU במאסטר; לטפל ידנית.",
        })

    # 3) Matcha — aggregated per Tom's instruction
    # MAT-500G: 30+21+3 = 54
    # MAT-18G (22-pack): 50+2 + 7 (50g tins folded in) = 59
    matcha_500 = 0
    for nm in [
        "Maruei Ceremonial Matcha Bag 500g",
        "Shizuoka Ceremonial Matcha Bag 500g",
        "Maruei XP Premium Ceremonial Matcha Bag 500g",
    ]:
        r = fc_by_batch_name.get(nm)
        if r:
            matcha_500 += float(r.get("אחר") or 0)

    matcha_18 = 0
    for nm in [
        "Shizuoka Ceremonial Matcha 22x18g.bags",
        "Maruei Ceremonial Matcha 22x18g.bags",
    ]:
        r = fc_by_batch_name.get(nm)
        if r:
            matcha_18 += float(r.get("אחר") or 0)
    # Fold 50g tins into the small-bag bucket per Tom
    for nm in [
        "Maruei Ceremonial Matcha Tin 50g",
        "Kogamo Organic Premium Ceremonial Matcha Tin 50g",
    ]:
        r = fc_by_batch_name.get(nm)
        if r:
            matcha_18 += float(r.get("אחר") or 0)

    if matcha_500 > 0:
        sku = "FG-MAT-500G"
        batches.append({
            "sheet_name": "MATCHA 500G",
            "hebrew_name": "מאצ'ה 500 גרם",
            "family_label": "מאצ'ה 500 גרם",
            "base_bom_id": None,
            "liquid": False,
            "pack_inputs": [{
                "label": "500g",
                "fill": 0,
                "default_units": matcha_500,
                "sku": sku,
                "pack_lines": master["per_sku"][sku]["pack"],
            }],
            "raws_per_l": {},
            "notes": "אגרגציה: Maruei + Shizuoka + Maruei XP, כולם מסוג 500 גרם.",
        })

    if matcha_18 > 0:
        # The Master BOM for FG-MAT-18G is per single 18g sachet, but Tom's
        # forecast unit is a 22-pack (one carton = 22 sachets). Rescale per-pack:
        # inner components × 22, carton × 1.
        sku = "FG-MAT-18G"
        raw_lines = master["per_sku"][sku]["pack"]
        adjusted = []
        for ln in raw_lines:
            cid = ln["comp_id"]
            if cid == "PKG-CARTON-MAT-22":
                # 0.0454545 × 22 = 1 (one carton per pack)
                adjusted.append({**ln, "qty": 1.0})
            else:
                adjusted.append({**ln, "qty": ln["qty"] * 22})
        batches.append({
            "sheet_name": "MATCHA 22x18G",
            "hebrew_name": "מאצ'ה 22×18 גרם",
            "family_label": "מאצ'ה 22×18 גרם",
            "base_bom_id": None,
            "liquid": False,
            "pack_inputs": [{
                "label": "חבילה של 22 שקיות",
                "fill": 0,
                "default_units": matcha_18,
                "sku": sku,
                "pack_lines": adjusted,
            }],
            "raws_per_l": {},
            "notes": "אגרגציה: כל המאצ'ה בחבילות הקטנות, כולל 7 פחיות 50 גרם שהוטמעו כאן. כל חבילה = 22 שקיות 18 גרם.",
        })

    return batches


# ────────────────────────────────────────────────────────────────────────────
# Excel rendering
# ────────────────────────────────────────────────────────────────────────────

# ── styling primitives
BLACK = "FF000000"
HEADER_BG = "FF1F2937"  # near-black
SECTION_BG = "FFF3F4F6"  # light gray
INPUT_BG = "FFFFF7CC"  # soft yellow
TITLE_FONT = Font(name="Arial", size=16, bold=True, color="FFFFFFFF")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="FF111827")
COL_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
DEFAULT_FONT = Font(name="Arial", size=10, color="FF111827")
INPUT_FONT = Font(name="Arial", size=11, bold=True, color="FF111827")
SUB_FONT = Font(name="Arial", size=9, italic=True, color="FF6B7280")

THIN = Side(style="thin", color="FFD1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

QTY_FMT_INT = '#,##0;[Red]-#,##0;"-"'
QTY_FMT_DEC = '#,##0.00;[Red]-#,##0.00;"-"'


def set_rtl(ws):
    ws.sheet_view.rightToLeft = True


def fill_cell(ws, coord, value=None, *, font=None, fill=None, align=None,
              border=BORDER, fmt=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border
    if fmt is not None:
        c.number_format = fmt
    return c


def header_fill(color=HEADER_BG):
    return PatternFill("solid", fgColor=color)


def section_fill():
    return PatternFill("solid", fgColor=SECTION_BG)


def input_fill():
    return PatternFill("solid", fgColor=INPUT_BG)


def merge(ws, range_str):
    ws.merge_cells(range_str)


def col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Per-batch sheet writer
# Layout (RTL — visually right-to-left, but cell coords are A,B,C,...):
#   Col A = label / row caption
#   Col B = input (units) / formula qty
#   Col C = code / UOM
#   Col D = supplier / extra
#   Col E = formula qty (for component lines, this is the canonical "qty to buy")
#
# We use a unified layout for component rows:
#   A: שם הרכיב | B: קוד | C: יח׳ | D: ספק | E: כמות
# Forecast rows:
#   A: גודל אריזה | B: כמות יחידות (input) | C: ליטר ליחידה | D: סה"כ ליטרים
#
# Component rows live in a known table. The summary uses SUMIFS on B (code) and E (qty).

def write_batch_sheet(wb, batch, summary_index):
    ws = wb.create_sheet(title=batch["sheet_name"])
    set_rtl(ws)
    col_widths(ws, [38, 15, 26, 12, 26])

    # Title bar
    ws.row_dimensions[1].height = 30
    ws["A1"] = batch["family_label"]
    fill_cell(ws, "A1", font=TITLE_FONT, fill=header_fill(),
              align=Alignment(horizontal="right", vertical="center"))
    merge(ws, "A1:E1")

    # Subline (small label "תוכנית רכש – מאי 2026")
    fill_cell(ws, "A2", value="תכנית רכש – מאי 2026",
              font=SUB_FONT, align=ALIGN_RIGHT, border=None)
    merge(ws, "A2:E2")

    # ── Forecast section
    row = 4
    fill_cell(ws, f"A{row}", value="תחזית מכירות",
              font=SECTION_FONT, fill=section_fill(), align=ALIGN_RIGHT)
    merge(ws, f"A{row}:E{row}")
    row += 1
    # column headers
    fill_cell(ws, f"A{row}", value="גודל אריזה", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"B{row}", value="כמות יחידות", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"C{row}", value="ליטר ליחידה", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"D{row}", value="ליטרים", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"E{row}", value="הערות", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    row += 1

    forecast_input_cells = []  # list of (cell_ref, fill_qty, label, sku, units_cell_for_pack_calc)
    for pi in batch["pack_inputs"]:
        fill_cell(ws, f"A{row}", value=pi["label"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
        # Input cell — editable
        units_cell = f"B{row}"
        fill_cell(ws, units_cell, value=pi["default_units"],
                  font=INPUT_FONT, fill=input_fill(), align=ALIGN_CENTER, fmt=QTY_FMT_INT)
        if batch["liquid"]:
            fill_cell(ws, f"C{row}", value=pi["fill"], font=DEFAULT_FONT,
                      align=ALIGN_CENTER, fmt='0.##" L"')
            fill_cell(ws, f"D{row}", value=f"=B{row}*C{row}",
                      font=DEFAULT_FONT, align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
        else:
            fill_cell(ws, f"C{row}", value="—", font=DEFAULT_FONT, align=ALIGN_CENTER)
            fill_cell(ws, f"D{row}", value="—", font=DEFAULT_FONT, align=ALIGN_CENTER)
        fill_cell(ws, f"E{row}", value="", font=DEFAULT_FONT, align=ALIGN_RIGHT)
        forecast_input_cells.append({
            "units_cell": units_cell,
            "fill_qty": pi["fill"],
            "label": pi["label"],
            "sku": pi["sku"],
            "pack_lines": pi["pack_lines"],
        })
        row += 1

    # totals row
    units_total_cell = f"B{row}"
    liters_total_cell = f"D{row}"
    fill_cell(ws, f"A{row}", value="סה״כ", font=Font(name="Arial", size=10, bold=True),
              fill=section_fill(), align=ALIGN_RIGHT)
    sum_units = "+".join(f["units_cell"] for f in forecast_input_cells)
    fill_cell(ws, units_total_cell, value=f"={sum_units}",
              font=Font(name="Arial", size=10, bold=True), fill=section_fill(),
              align=ALIGN_CENTER, fmt=QTY_FMT_INT)
    fill_cell(ws, f"C{row}", value="", fill=section_fill())
    if batch["liquid"]:
        sum_liters = "+".join(f"{f['units_cell']}*{f['fill_qty']}" for f in forecast_input_cells)
        fill_cell(ws, liters_total_cell, value=f"={sum_liters}",
                  font=Font(name="Arial", size=10, bold=True),
                  fill=section_fill(), align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
    else:
        fill_cell(ws, liters_total_cell, value="", fill=section_fill())
    fill_cell(ws, f"E{row}", value="", fill=section_fill())
    row += 2

    # ── Packaging section
    fill_cell(ws, f"A{row}", value="אריזות לרכוש",
              font=SECTION_FONT, fill=section_fill(), align=ALIGN_RIGHT)
    merge(ws, f"A{row}:E{row}")
    row += 1
    fill_cell(ws, f"A{row}", value="רכיב", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"B{row}", value="קוד", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"C{row}", value="ספק", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"D{row}", value="יח'", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"E{row}", value="כמות לרכישה", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
    row += 1

    # Each pack line is a per-unit BOM line for one specific pack-size SKU.
    # We aggregate by component_id across pack-sizes to produce ONE row per component.
    # Formula sums (units_of_size * qty_per_unit) across the sizes that include this comp.
    pack_agg = OrderedDict()  # comp_id → {name, uom, supplier, terms: [(units_cell, qty_per_unit)]}
    for fi in forecast_input_cells:
        for line in fi["pack_lines"]:
            cid = line["comp_id"]
            agg = pack_agg.get(cid)
            if not agg:
                pack_agg[cid] = {
                    "name": line["name"],
                    "uom": line["uom"],
                    "supplier": line["supplier"],
                    "terms": [(fi["units_cell"], line["qty"])],
                }
            else:
                agg["terms"].append((fi["units_cell"], line["qty"]))

    for cid, info in pack_agg.items():
        fill_cell(ws, f"A{row}", value=info["name"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
        fill_cell(ws, f"B{row}", value=cid, font=DEFAULT_FONT, align=ALIGN_CENTER)
        fill_cell(ws, f"C{row}", value=info["supplier"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
        fill_cell(ws, f"D{row}", value=info["uom"], font=DEFAULT_FONT, align=ALIGN_CENTER)
        terms = "+".join(
            f"{u}*{q}" if abs(q - 1) > 1e-9 else f"{u}"
            for u, q in info["terms"]
        )
        # Detect integer vs fractional → format accordingly
        max_frac = any(abs(q - round(q)) > 1e-9 for _, q in info["terms"])
        fmt = QTY_FMT_DEC if max_frac else QTY_FMT_INT
        fill_cell(ws, f"E{row}", value=f"={terms}", font=DEFAULT_FONT,
                  align=ALIGN_CENTER, fmt=fmt)
        # Track for summary: classify by component prefix, not by section,
        # because matcha pack-BOMs include RAW-MATCHA-BULK as a "pack" line.
        if cid.startswith("RAW-"):
            kind = "raw"
        elif cid.startswith("PKG-FILTER"):
            kind = "filter"
        else:
            kind = "pack"
        summary_index.setdefault(cid, {
            "name": info["name"],
            "uom": info["uom"],
            "supplier": info["supplier"],
            "kind": kind,
        })
        row += 1

    # ── Raw materials section (only for liquid batches)
    if batch["liquid"] and batch["raws_per_l"]:
        row += 1
        fill_cell(ws, f"A{row}", value="חומרי גלם לרכוש (לפי סה״כ ליטרים)",
                  font=SECTION_FONT, fill=section_fill(), align=ALIGN_RIGHT)
        merge(ws, f"A{row}:E{row}")
        row += 1
        fill_cell(ws, f"A{row}", value="רכיב", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"B{row}", value="קוד", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"C{row}", value="ספק", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"D{row}", value="יח'", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"E{row}", value="כמות לרכישה", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        row += 1

        for cid, info in batch["raws_per_l"].items():
            fill_cell(ws, f"A{row}", value=info["name"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"B{row}", value=cid, font=DEFAULT_FONT, align=ALIGN_CENTER)
            fill_cell(ws, f"C{row}", value=info["supplier"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"D{row}", value=info["uom"], font=DEFAULT_FONT, align=ALIGN_CENTER)
            ratio = info["qty_per_l"]
            # Round ratio to 6 decimals to keep formula clean
            ratio_str = f"{ratio:.6f}".rstrip("0").rstrip(".")
            if ratio_str == "" or ratio_str == "-":
                ratio_str = "0"
            fill_cell(ws, f"E{row}", value=f"={liters_total_cell}*{ratio_str}",
                      font=DEFAULT_FONT, align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
            kind = "filter" if cid.startswith("PKG-FILTER") else "raw"
            summary_index.setdefault(cid, {
                "name": info["name"],
                "uom": info["uom"],
                "supplier": info["supplier"],
                "kind": kind,
            })
            row += 1

    # Notes
    if batch["notes"]:
        row += 1
        fill_cell(ws, f"A{row}", value="הערה", font=Font(name="Arial", size=10, bold=True),
                  align=ALIGN_RIGHT, border=None)
        fill_cell(ws, f"B{row}", value=batch["notes"], font=DEFAULT_FONT,
                  align=ALIGN_RIGHT, border=None)
        merge(ws, f"B{row}:E{row}")

    # Freeze top
    ws.freeze_panes = "A4"
    return ws


# ── Summary sheet
def write_summary_sheet(wb, batches, summary_index):
    ws = wb.create_sheet(title="סיכום רכש", index=0)
    set_rtl(ws)
    col_widths(ws, [38, 16, 26, 10, 16, 14])

    ws.row_dimensions[1].height = 32
    fill_cell(ws, "A1", value="סיכום רכש כולל — מאי 2026",
              font=TITLE_FONT, fill=header_fill(),
              align=Alignment(horizontal="right", vertical="center"))
    merge(ws, "A1:F1")
    fill_cell(ws, "A2", value=f"מקובץ מ-{len(batches)} מוצרים | עדכון אוטומטי לפי שינוי בתחזית בכל גליון",
              font=SUB_FONT, align=ALIGN_RIGHT, border=None)
    merge(ws, "A2:F2")

    row = 4
    # Group by kind: packaging, raw, filter
    sections = [
        ("אריזות", "pack"),
        ("חומרי גלם", "raw"),
        ("פילטרים וחומרי עזר לייצור", "filter"),
    ]

    sheet_names = [b["sheet_name"] for b in batches]

    for sec_label, sec_kind in sections:
        items = [(cid, info) for cid, info in summary_index.items() if info["kind"] == sec_kind]
        if not items:
            continue
        # section header
        fill_cell(ws, f"A{row}", value=sec_label,
                  font=Font(name="Arial", size=12, bold=True, color="FFFFFFFF"),
                  fill=header_fill(),
                  align=Alignment(horizontal="right", vertical="center"))
        merge(ws, f"A{row}:F{row}")
        row += 1

        # column headers
        fill_cell(ws, f"A{row}", value="רכיב", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"B{row}", value="קוד", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"C{row}", value="ספק", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"D{row}", value="יח'", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"E{row}", value="סה״כ לרכישה", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        fill_cell(ws, f"F{row}", value="מספר מוצרים", font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        row += 1

        # Sort by name (raw/filter) or supplier+name (pack)
        if sec_kind == "pack":
            items.sort(key=lambda x: (x[1]["supplier"] or "", x[1]["name"] or ""))
        else:
            items.sort(key=lambda x: x[1]["name"] or "")

        for cid, info in items:
            fill_cell(ws, f"A{row}", value=info["name"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"B{row}", value=cid, font=DEFAULT_FONT, align=ALIGN_CENTER)
            fill_cell(ws, f"C{row}", value=info["supplier"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"D{row}", value=info["uom"], font=DEFAULT_FONT, align=ALIGN_CENTER)

            # Total qty: sum across sheets via SUMIFS lookup (qty in col E, code in col B)
            terms = []
            for sn in sheet_names:
                terms.append(f"IFERROR(SUMIFS('{sn}'!E:E,'{sn}'!B:B,B{row}),0)")
            formula = "=" + "+".join(terms)
            fill_cell(ws, f"E{row}", value=formula, font=DEFAULT_FONT,
                      align=ALIGN_CENTER, fmt=QTY_FMT_DEC)

            # Number of products contributing (count sheets where SUMIFS > 0)
            count_terms = []
            for sn in sheet_names:
                count_terms.append(f"IF(IFERROR(SUMIFS('{sn}'!E:E,'{sn}'!B:B,B{row}),0)>0,1,0)")
            count_formula = "=" + "+".join(count_terms)
            fill_cell(ws, f"F{row}", value=count_formula, font=DEFAULT_FONT,
                      align=ALIGN_CENTER, fmt='0')
            row += 1

        row += 1  # spacer

    # KPIs at top
    # Move down two rows for the totals block
    fill_cell(ws, f"A3", value=f"סה״כ פריטים שונים לרכישה: {len(summary_index)}",
              font=Font(name="Arial", size=10, bold=True), align=ALIGN_RIGHT, border=None)

    ws.freeze_panes = "A5"


# ────────────────────────────────────────────────────────────────────────────
# Index sheet (cover)
# ────────────────────────────────────────────────────────────────────────────
def write_index_sheet(wb, batches):
    ws = wb.create_sheet(title="פתיחה", index=0)
    set_rtl(ws)
    col_widths(ws, [38, 18, 18, 32])
    ws.row_dimensions[1].height = 36
    fill_cell(ws, "A1", value="תכנית רכש — מאי 2026",
              font=TITLE_FONT, fill=header_fill(),
              align=Alignment(horizontal="right", vertical="center"))
    merge(ws, "A1:D1")
    fill_cell(ws, "A2",
              value="גליון פתיחה | תחזית מכירות בכל מוצר ניתנת לעריכה — נוסחאות מתעדכנות אוטומטית.",
              font=SUB_FONT, align=ALIGN_RIGHT, border=None)
    merge(ws, "A2:D2")

    row = 4
    fill_cell(ws, f"A{row}", value="מוצר", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"B{row}", value="סה״כ יחידות", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"C{row}", value="סה״כ ליטרים", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    fill_cell(ws, f"D{row}", value="הערות", font=COL_HEADER_FONT,
              fill=header_fill(), align=ALIGN_CENTER)
    row += 1
    grand_units_terms = []
    grand_liters_terms = []
    for b in batches:
        sn = b["sheet_name"]
        # Find the units and liters total cells in each sheet
        # By construction: forecast input rows start at row 6, totals row at 6 + len(pack_inputs)
        n_inputs = len(b["pack_inputs"])
        totals_row = 6 + n_inputs  # row 5 is column header; first input at 6
        units_total = f"'{sn}'!B{totals_row}"
        liters_total = f"'{sn}'!D{totals_row}" if b["liquid"] else None

        fill_cell(ws, f"A{row}", value=b["family_label"], font=DEFAULT_FONT, align=ALIGN_RIGHT)
        fill_cell(ws, f"B{row}", value=f"={units_total}",
                  font=DEFAULT_FONT, align=ALIGN_CENTER, fmt=QTY_FMT_INT)
        if liters_total:
            fill_cell(ws, f"C{row}", value=f"={liters_total}",
                      font=DEFAULT_FONT, align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
            grand_liters_terms.append(liters_total)
        else:
            fill_cell(ws, f"C{row}", value="—", font=DEFAULT_FONT, align=ALIGN_CENTER)
        grand_units_terms.append(units_total)
        fill_cell(ws, f"D{row}", value=b.get("notes", "") or "",
                  font=SUB_FONT, align=ALIGN_RIGHT)
        row += 1

    # grand totals
    fill_cell(ws, f"A{row}", value="סה״כ",
              font=Font(name="Arial", size=10, bold=True),
              fill=section_fill(), align=ALIGN_RIGHT)
    fill_cell(ws, f"B{row}", value=f"={'+'.join(grand_units_terms)}",
              font=Font(name="Arial", size=10, bold=True),
              fill=section_fill(), align=ALIGN_CENTER, fmt=QTY_FMT_INT)
    if grand_liters_terms:
        fill_cell(ws, f"C{row}", value=f"={'+'.join(grand_liters_terms)}",
                  font=Font(name="Arial", size=10, bold=True),
                  fill=section_fill(), align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
    fill_cell(ws, f"D{row}", value="", fill=section_fill())

    ws.freeze_panes = "A5"


# ────────────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────────────
def compute_demand(batches, summary_index):
    """Mirror Excel's formulas in Python to get a numeric per-component demand."""
    totals = OrderedDict()  # comp_id → {name, supplier, uom, kind, total}

    def bump(cid, qty, default_info):
        if qty <= 0:
            return
        if cid not in totals:
            si = summary_index.get(cid, {})
            kind = si.get("kind") or default_info.get("kind", "raw")
            totals[cid] = {
                "name": si.get("name") or default_info.get("name", ""),
                "supplier": si.get("supplier") or default_info.get("supplier", ""),
                "uom": si.get("uom") or default_info.get("uom", ""),
                "kind": kind,
                "total": 0.0,
            }
        totals[cid]["total"] += qty

    for b in batches:
        # Liters total (for raws)
        if b["liquid"]:
            total_l = sum(p["default_units"] * p["fill"] for p in b["pack_inputs"])
        else:
            total_l = 0
        # Pack lines aggregated across pack-sizes for this batch
        pack_agg = {}  # cid → (qty, info)
        for pi in b["pack_inputs"]:
            for ln in pi["pack_lines"]:
                cid = ln["comp_id"]
                add = pi["default_units"] * ln["qty"]
                if cid in pack_agg:
                    pack_agg[cid] = (pack_agg[cid][0] + add, pack_agg[cid][1])
                else:
                    pack_agg[cid] = (add, ln)
        for cid, (qty, ln) in pack_agg.items():
            if cid.startswith("RAW-"):
                kind = "raw"
            elif cid.startswith("PKG-FILTER"):
                kind = "filter"
            else:
                kind = "pack"
            bump(cid, qty, {"name": ln["name"], "supplier": ln["supplier"],
                            "uom": ln["uom"], "kind": kind})
        # Raws
        for cid, info in b.get("raws_per_l", {}).items():
            qty = total_l * info["qty_per_l"]
            kind = "filter" if cid.startswith("PKG-FILTER") else "raw"
            bump(cid, qty, {"name": info["name"], "supplier": info["supplier"],
                            "uom": info["uom"], "kind": kind})
    return totals


def write_stock_template(totals):
    """Single-sheet stand-alone template — Tom fills 'מלאי קיים' col."""
    wb = Workbook()
    ws = wb.active
    ws.title = "מלאי ורכש מאי 2026"
    set_rtl(ws)
    col_widths(ws, [38, 28, 28, 8, 16, 16, 16])

    ws.row_dimensions[1].height = 32
    fill_cell(ws, "A1", value="תכנית רכש מאי 2026 — מילוי מלאי",
              font=TITLE_FONT, fill=header_fill(),
              align=Alignment(horizontal="right", vertical="center"))
    merge(ws, "A1:G1")
    fill_cell(
        ws, "A2",
        value="מלא את עמודת 'מלאי קיים' בכמות שיש לך עכשיו (יח׳/ק״ג/L לפי הטור 'יח'). "
              "עמודת 'לקנייה' תתעדכן אוטומטית — הכל בנוסחה. כולל מתכון NAMASTEA המעודכן.",
        font=SUB_FONT, align=ALIGN_RIGHT, border=None)
    merge(ws, "A2:G2")

    sections = [
        ("אריזות", "pack"),
        ("חומרי גלם", "raw"),
        ("פילטרים וחומרי עזר לייצור", "filter"),
    ]

    row = 4
    for sec_label, sec_kind in sections:
        items = [(cid, info) for cid, info in totals.items() if info["kind"] == sec_kind]
        if not items:
            continue
        # Match the original summary sheet ordering exactly:
        # pack → by supplier+name; raw/filter → by name only.
        if sec_kind == "pack":
            items.sort(key=lambda x: ((x[1]["supplier"] or ""), (x[1]["name"] or "")))
        else:
            items.sort(key=lambda x: x[1]["name"] or "")

        # section header
        fill_cell(ws, f"A{row}", value=sec_label,
                  font=Font(name="Arial", size=12, bold=True, color="FFFFFFFF"),
                  fill=header_fill(),
                  align=Alignment(horizontal="right", vertical="center"))
        merge(ws, f"A{row}:G{row}")
        row += 1

        # column headers
        for col, label in zip("ABCDEFG", [
            "רכיב", "קוד", "ספק", "יח'",
            "דרישה (מאי)", "מלאי קיים", "לקנייה",
        ]):
            fill_cell(ws, f"{col}{row}", value=label,
                      font=COL_HEADER_FONT, fill=header_fill(), align=ALIGN_CENTER)
        row += 1

        for cid, info in items:
            fill_cell(ws, f"A{row}", value=info["name"],
                      font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"B{row}", value=cid,
                      font=DEFAULT_FONT, align=ALIGN_CENTER)
            fill_cell(ws, f"C{row}", value=info["supplier"],
                      font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"D{row}", value=info["uom"],
                      font=DEFAULT_FONT, align=ALIGN_CENTER)
            # demand
            qty = info["total"]
            fill_cell(ws, f"E{row}", value=round(qty, 2),
                      font=DEFAULT_FONT, align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
            # stock — editable, default 0
            fill_cell(ws, f"F{row}", value=0,
                      font=INPUT_FONT, fill=input_fill(),
                      align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
            # to-buy = MAX(0, demand - stock)
            fill_cell(ws, f"G{row}", value=f"=MAX(0,E{row}-F{row})",
                      font=Font(name="Arial", size=10, bold=True, color="FFB91C1C"),
                      align=ALIGN_CENTER, fmt=QTY_FMT_DEC)
            row += 1

        row += 1  # spacer between sections

    # Top counters
    fill_cell(ws, "A3", value=f"סה״כ פריטים שונים בתכנית: {len(totals)}",
              font=Font(name="Arial", size=10, bold=True),
              align=ALIGN_RIGHT, border=None)

    ws.freeze_panes = "A5"
    print(f"Saving stock template -> {STOCK_TEMPLATE_FILE}")
    wb.save(STOCK_TEMPLATE_FILE)


def main():
    print("Loading master data…")
    master = load_master()
    apply_recipe_overrides(master)
    print(f"  items={len(master['items'])}  base_recipes={len(master['base_recipes'])}"
          f"  (NAMASTEA recipe override applied)")

    print("Loading forecast…")
    forecast_rows = load_forecast()
    print(f"  rows={len(forecast_rows)}")

    print("Building scope…")
    batches = build_scope(forecast_rows, master)
    print(f"  batches in scope: {len(batches)}")
    for b in batches:
        units = sum(p["default_units"] for p in b["pack_inputs"])
        liters = sum(p["default_units"] * p["fill"] for p in b["pack_inputs"])
        print(f"   - {b['sheet_name']}: {int(units)} units / {liters:.1f} L"
              f" / {len(b['pack_inputs'])} pack sizes / "
              f"{len(b['raws_per_l'])} raws")

    print("Rendering workbook…")
    wb = Workbook()
    wb.remove(wb.active)  # drop default

    # We'll build batches first to populate summary_index, then write summary at index 0
    summary_index = OrderedDict()
    for b in batches:
        write_batch_sheet(wb, b, summary_index)

    write_summary_sheet(wb, batches, summary_index)
    write_index_sheet(wb, batches)

    print(f"  unique components in summary: {len(summary_index)}")

    print(f"Saving -> {OUT_FILE}")
    try:
        wb.save(OUT_FILE)
    except PermissionError:
        print(f"  (skipped — file is open in Excel; close it to regenerate the full plan)")

    # Always emit the stand-alone stock-fill template
    totals = compute_demand(batches, summary_index)
    try:
        write_stock_template(totals)
    except PermissionError:
        print(f"  (skipped — stock template is open in Excel)")
    print("Done.")


if __name__ == "__main__":
    main()
