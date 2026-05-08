"""
COP Sangria Sync — 2026-05-07
Tom-locked: cost-of-production August 2025 is source of truth for white sangria
(Babka 1L, Elita 0.3L/0.75L, 3.85L jerrycan), pink sangria 1L, and the Fresh+Calm
sub-bases that feed into them.

All five BASE BOMs are normalized to a 500 L batch result with COP-derived ratios.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import shutil
import sys, io
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WB_PATH = "GT_Factory_OS.xlsx"

NEW_VERSION = "V5_COP_AUG2025_2026-05-07"
NEW_NOTES = (
    "DYNAMIC_RATIO_MODEL — 500 L canonical batch. "
    "Lines synced 2026-05-07 from cost of production August 2025 (Tom-locked)."
)

# (line_no -> (final_qty_for_500L, qty_per_L, optional_status_override))
PINK = {
    1: (210.0, 0.420, None),
    2: (290.0, 0.580, None),
    3: (1.10,  0.0022, None),
    4: (0.85,  0.0017, None),
}
WHI_BABKA = {
    1: (175.0, 0.350, None),
    2: (325.0, 0.650, None),
    3: (1.20,  0.0024, None),
    4: (0.95,  0.0019, None),
}
WHI_ELITA = {
    1: (75.0,  0.150, None),
    2: (365.0, 0.730, None),
    3: (0.50,  0.0010, None),
    4: (0.0,   0.0,   "RETIRED"),  # preservative removed per COP
    5: (30.0,  0.060, None),
    6: (15.0,  0.030, None),
    7: (15.0,  0.030, None),
}
FRESH = {
    1: (30.0,   0.060,  None),
    2: (1.0,    0.002,  None),
    3: (10.75,  0.0215, None),
    4: (190.0,  0.380,  None),
    5: (430.0,  0.860,  None),
    6: (1.0,    0.002,  None),
    7: (1.0,    0.002,  None),
}
CALM = {
    1: (21.5,   0.043,  None),
    2: (30.0,   0.060,  None),
    3: (12.0,   0.024,  None),
    4: (1.7,    0.0034, None),
    5: (210.0,  0.420,  None),
    6: (504.0,  1.008,  None),
    7: (1.2,    0.0024, None),
    8: (1.0,    0.002,  None),
    9: (1.0,    0.002,  None),
}
# new line to append for Calm: Lime Puree
CALM_NEW_LINE = {
    "LINE_KEY": "BOM-BASE-CAL-REG-L10",
    "BOM_ID": "BOM-BASE-CAL-REG",
    "LINE_NO": 10,
    "BOM_KIND": "BASE",
    "DISPLAY_FAMILY": "CALM",
    "SWEETNESS": "REGULAR",
    "PACK_SIZE": None,
    "COMPONENT_REF_TYPE": "RAW_NAME",
    "FINAL_COMPONENT_ID": "RAW-LIME-PUREE",
    "FINAL_COMPONENT_NAME": "Lime Puree (Ristretto)",
    "FINAL_COMPONENT_QTY": 4.8,
    "COMPONENT_UOM": "KG",
    "STATUS": "ACTIVE",
    "LINE_CHECK": "OK",
    "OWNER_NOTES": "Added 2026-05-07 from cost of production August 2025 (Calm run 2).",
    "STD_COST_PER_UOM": None,
    "LINE_STD_COST": None,
    "SCALING_METHOD": "RATIO",
    "QTY_PER_L_OUTPUT": 0.0096,
}

PLAN = [
    # (bom_id, header_row, line_rows[line_no -> sheet_row], updates, label)
    ("BOM-BASE-SAN-PIN-REG",     22, {1:164, 2:165, 3:166, 4:167},                                       PINK,        "Pink Sangria base"),
    ("BOM-BASE-SAN-WHI-REG",     25, {1:183, 2:184, 3:185, 4:186},                                       WHI_BABKA,   "White Sangria 1L (Babka) base"),
    ("BOM-BASE-SAN-WHI-ELI-REG", 24, {1:176, 2:177, 3:178, 4:179, 5:180, 6:181, 7:182},                  WHI_ELITA,   "White Sangria Elita base"),
    ("BOM-BASE-FRE-REG",         18, {1:121, 2:122, 3:123, 4:124, 5:125, 6:126, 7:127},                  FRESH,       "Fresh base"),
    ("BOM-BASE-CAL-REG",          9, {1:31, 2:32, 3:33, 4:34, 5:35, 6:36, 7:37, 8:38, 9:39},             CALM,        "Calm base"),
]

COL = {
    "LINE_KEY": 1, "BOM_ID": 2, "LINE_NO": 3, "BOM_KIND": 4, "DISPLAY_FAMILY": 5,
    "SWEETNESS": 6, "PACK_SIZE": 7, "COMPONENT_REF_TYPE": 8, "FINAL_COMPONENT_ID": 9,
    "FINAL_COMPONENT_NAME": 10, "FINAL_COMPONENT_QTY": 11, "COMPONENT_UOM": 12,
    "STATUS": 13, "LINE_CHECK": 14, "OWNER_NOTES": 15, "STD_COST_PER_UOM": 16,
    "LINE_STD_COST": 17, "SCALING_METHOD": 18, "QTY_PER_L_OUTPUT": 19,
}

def main():
    wb = openpyxl.load_workbook(WB_PATH)
    ws_h = wb["BOM_HEADER"]
    ws_l = wb["BOM_LINES"]

    print(f"\n=== COP Sangria Sync — applying changes ===\n")

    for bom_id, hdr_row, line_rows, updates, label in PLAN:
        # Verify header BOM_ID
        actual = ws_h.cell(hdr_row, 1).value
        assert actual == bom_id, f"Header row {hdr_row} expected {bom_id}, got {actual}"

        # Update header batch size (col 10), version (col 12), notes (col 15)
        old_batch = ws_h.cell(hdr_row, 10).value
        ws_h.cell(hdr_row, 10).value = 500
        ws_h.cell(hdr_row, 11).value = "L"
        old_ver = ws_h.cell(hdr_row, 12).value
        ws_h.cell(hdr_row, 12).value = NEW_VERSION
        ws_h.cell(hdr_row, 15).value = NEW_NOTES

        print(f"--- {label}  ({bom_id}) ---")
        print(f"  HEADER: batch {old_batch} L → 500 L | version {old_ver} → {NEW_VERSION}")

        # Update each line
        for line_no, (qty, per_l, status_override) in updates.items():
            row = line_rows[line_no]
            actual_bid = ws_l.cell(row, COL["BOM_ID"]).value
            actual_ln = ws_l.cell(row, COL["LINE_NO"]).value
            assert actual_bid == bom_id, f"Row {row} BOM_ID expected {bom_id}, got {actual_bid}"
            assert int(actual_ln) == line_no, f"Row {row} LINE_NO expected {line_no}, got {actual_ln!r}"

            cname = ws_l.cell(row, COL["FINAL_COMPONENT_NAME"]).value or ws_l.cell(row, COL["FINAL_COMPONENT_ID"]).value
            old_qty = ws_l.cell(row, COL["FINAL_COMPONENT_QTY"]).value
            old_status = ws_l.cell(row, COL["STATUS"]).value
            ws_l.cell(row, COL["FINAL_COMPONENT_QTY"]).value = qty
            ws_l.cell(row, COL["QTY_PER_L_OUTPUT"]).value = per_l
            ws_l.cell(row, COL["SCALING_METHOD"]).value = "RATIO"
            if status_override:
                ws_l.cell(row, COL["STATUS"]).value = status_override
                ws_l.cell(row, COL["OWNER_NOTES"]).value = (
                    "Retired 2026-05-07: not present in cost of production August 2025."
                )
            new_status = ws_l.cell(row, COL["STATUS"]).value
            arrow = f"{old_qty} → {qty}"
            extra = f"  [STATUS {old_status} → {new_status}]" if status_override else ""
            print(f"  L{line_no} {cname}: {arrow}  per_L={per_l}{extra}")

    # Append new line for Calm (Lime Puree)
    target_row = None
    for r in range(6, ws_l.max_row + 2):
        if ws_l.cell(r, 1).value is None and ws_l.cell(r, 2).value is None:
            target_row = r
            break
    print(f"\n--- Calm — append Lime Puree (new line) ---")
    print(f"  appending at row {target_row}")
    for field, val in CALM_NEW_LINE.items():
        ws_l.cell(target_row, COL[field]).value = val
    print(f"  L10 RAW-LIME-PUREE (Lime Puree (Ristretto)): 4.8 KG  per_L=0.0096")

    wb.save(WB_PATH)
    print(f"\n✓ Saved to {WB_PATH}")

if __name__ == "__main__":
    main()
