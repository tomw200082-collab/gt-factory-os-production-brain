"""
Build the final May 2026 purchase order Excel from:
  - Filled stock template (to-buy quantities)
  - Prices extracted from invoice images in 'חשבוניות ספקים'

Layout:
  - Single sheet, Hebrew RTL.
  - Columns: רכיב | קוד | ספק | יח׳ | כמות לקנייה | מחיר ליחידה | עלות שורה
  - Row groups by supplier; each group has a subtotal.
  - Empty 'מחיר ליחידה' cells are highlighted yellow → user fills them in
    and 'עלות שורה' / supplier subtotal / grand total recompute via formula.
"""

import os
import math
from collections import defaultdict, OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STOCK_TEMPLATE = os.path.join(ROOT, "GT_May_2026_Stock_Fill_Template.xlsx")
OUT_FILE = os.path.join(ROOT, "GT_May_2026_Final_Purchase_Order.xlsx")


# ────────────────────────────────────────────────────────────────────────────
# Price catalog extracted from invoice images
# ────────────────────────────────────────────────────────────────────────────
# ALL prices are PRE-VAT (net) where extractable.
# Unit = price per 1 KG / 1 L / 1 UNIT (always per 1 of inventory UOM).
# notes: "src=<file>", "date=YYYY-MM-DD", or 'estimated' / 'most-recent-of-N'.
PRICES = {
    # ── AMITEA — invoices are generic "תה במבחר" lump sums; per-kg price not extractable
    "RAW-PUER":               {"price": None, "note": "מחיר לא חולץ — חשבוניות AMITEA סוכמות בלי פירוט פר תה. הזן ידנית."},
    "RAW-SENCHA":             {"price": None, "note": "אותו מצב כמו פו-אר. הזן ידנית."},
    "RAW-BLACK-TEA":          {"price": None, "note": "אותו מצב — AMITEA סוכם, ללא פירוט. הזן ידנית."},

    # ── Ristretto (puree)
    "RAW-LEMON-PUREE":        {"price": 28.00,   "note": "Ristretto, מחית לימון 5kg @ 140/שקית = 28/kg. חשבונית 09/04/2026."},
    "RAW-LIME-PUREE":         {"price": 34.20,   "note": "Ristretto, מחית ליים 1kg @ 34.20/kg. חשבוניות עקביות."},
    "RAW-LYCHEE-PUREE":       {"price": 40.00,   "note": "Ristretto, מחית ליצ'י 1kg @ 40/kg."},
    "RAW-PASSION-FRUIT-PUREE":{"price": 34.97,   "note": "Ristretto, מחית פסיפלורה 5kg @ 174.84/שקית = 34.97/kg."},

    # ── Tempo (alcohol & wine)
    "RAW-WINE-RED":           {"price": 12.64,   "note": "Tempo, SYMPHONY אדום 23L @ 290.76/ג'ריקן = 12.64/L. חשבונית 26/02/2026."},
    "RAW-WINE-WHITE":         {"price": 12.64,   "note": "Tempo, SYMPHONY לבן 23L @ 290.75/ג'ריקן = 12.64/L."},
    "RAW-RUM":                {"price": 131.90,  "note": "Tempo, רום הוואנה קלאב 0.7L @ 92.33/בקבוק = 131.90/L. חשבונית 08/01/2026."},
    "RAW-AMARETTO":           {"price": None,    "note": "לא נמצאה חשבונית של אמרטו. הזן ידנית."},

    # ── Habshush Arie / Tavlinei Bar (spices)
    "RAW-HIBISCUS":           {"price": 25.80,   "note": "Habshush Arie 14/04/2026 @ 25.80/kg."},
    "RAW-LUISA":              {"price": 86.00,   "note": "Habshush Arie 05/03/2026 @ 86/kg."},
    "RAW-EL":                 {"price": 218.00,  "note": "Habshush Arie 05/03/2026 (הל ירוק) @ 218/kg. (BAR ישן 145/kg.)"},
    "RAW-GINGER-CRUSHED":     {"price": 34.00,   "note": "Habshush Arie 05/03/2026 @ 34/kg."},
    "RAW-PAPER":              {"price": 43.00,   "note": "Habshush Arie 05/03/2026 (פלפל שחור שלם) @ 43/kg."},
    "RAW-CINNAMON":           {"price": 22.00,   "note": "Habshush Arie 05/03/2026 (קינמון סיני) @ 22/kg."},
    "RAW-NANA":               {"price": 24.00,   "note": "Habshush Arie 12/04/2026 (נענע ספרמינט) @ 24/kg."},
    "RAW-CLOVE":              {"price": 54.00,   "note": "BAR Spices 10/08/2025 @ 54/kg. (לא מצאתי חשבונית טריה יותר.)"},
    "RAW-MELISA":             {"price": None,    "note": "לא נמצאה במחירונים. הזן ידנית."},
    "RAW-DRIED-ORANGE":       {"price": 68.00,   "note": "Habshush Arie 12/04/2026 @ 68/kg."},
    "RAW-DRIED-LEMON":        {"price": None,    "note": "לא נמצאה חשבונית ספציפית של לימון יבש. הזן ידנית."},
    "RAW-APPLE-DRY":          {"price": 66.00,   "note": "Habshush Arie 11/02/2026 (תפוח עץ טבעי) @ 66/kg."},

    # ── Tsabar Arizot (bottles + caps)
    "PKG-BOTTLE-500ML":       {"price": 1.82,    "note": "Tsabar Arizot 29/03/2026, בקבוק 500ml חום פיה 28 @ 1.82/יח'."},
    "PKG-BOTTLE-750ML-GREEN": {"price": None,    "note": "לא נמצאה חשבונית של בקבוק 750ml ירוק. הזן ידנית."},
    "PKG-CAP-BLACK-METAL-28": {"price": None,    "note": "Tsabar 25/02/26 הציג רק פקק זהב @ 0.20. אין מחיר עדכני לפקק שחור — הזן ידנית."},

    # ── Rozenfeld (1L bottles)
    "PKG-BOTTLE-1L":          {"price": 2.54,    "note": "Rosenfeld Packaging 24/02/2026, בקבוק 1L אמבר @ 2.54/יח'."},

    # ── Eliran (cartons) — two carton sizes seen, mapping to specific BOM not 100% certain
    "PKG-CARTON-1L":          {"price": 3.98,    "note": "אלירן קרטונים, קרטון 36×26×31 ס\"מ @ 3.98/יח' (15/03/2026, 03/12/2025)."},
    "PKG-CARTON-500ML":       {"price": 2.92,    "note": "אלירן קרטונים, קרטון מודפס לבן 30×20.5×22.5 ס\"מ @ 2.92/יח' (15/04/2026). יתכן שזה בעצם 1L, לאמת מול הקרטון בפועל."},
    "PKG-CARTON-750ML":       {"price": None,    "note": "לא נמצאה חשבונית ספציפית. הזן ידנית."},
    "PKG-CARTON-MAT-22":      {"price": None,    "note": "לא ברור איזו מהקרטונים של אלירן הוא MAT-22. הזן ידנית."},

    # ── Miki labels
    "PKG-LABEL-MAT-18G":      {"price": 0.40,    "note": "מיקי מדבקות 11/02/2026, מדבקת מאצ'ה 60×40 מ\"מ @ 0.40/יח'."},
    "PKG-LABEL-MAT-500G":     {"price": 0.40,    "note": "מיקי מדבקות 11/02/2026, מדבקת מאצ'ה עגולה 120 מ\"מ @ 0.40/יח'."},
    "PKG-LABEL-MAT-100G":     {"price": 0.40,    "note": "אותו מחיר."},
    "PKG-LABEL-MAT-30G":      {"price": 0.40,    "note": "אותו מחיר."},
    "PKG-LABEL-SAN-RED-750ML":{"price": None,    "note": "לא נמצאה חשבונית ספציפית של מדבקת סנגריה אדומה. הזן ידנית."},

    # ── Filters
    "PKG-FILTER-100MICRON":   {"price": None,    "note": "לא נמצאה חשבונית של ספק הפילטרים. הזן ידנית."},
    "PKG-FILTER-200MICRON":   {"price": None,    "note": "אותו מצב — לא נמצאה חשבונית. הזן ידנית."},

    # ── Propack (matcha bags)
    "PKG-BAG-MAT-500G":       {"price": None,    "note": "Propack — חשבוניות הציגו רק שקית 9×13 שחור (ה-18g), לא 500g. הזן ידנית."},
    "PKG-BAG-MAT-18G":        {"price": 0.25,    "note": "Propack 23/04/2026, שקית אלומיניום 9×13 שחור @ 0.25/יח'."},
    "PKG-BAG-MAT-100G":       {"price": None,    "note": "לא נמצאה חשבונית. הזן ידנית."},

    # ── Ziv Chemicals
    "RAW-STABILISER":         {"price": 75.00,   "note": "Ziv Chemicals 15/01/2026, קסטנגאן (E415) @ 75/kg."},
    "RAW-PRESERVATIVE":       {"price": None,    "note": "לא נמצאה חשבונית של חומר משמר ב-bulk. הזן ידנית."},
    "RAW-LEMON-ACID":         {"price": None,    "note": "לא נמצאה חשבונית של חומצת לימון. הזן ידנית."},

    # ── Matcha bulk
    "RAW-MATCHA-BULK":        {"price": None,    "note": "מחיר מאצ'ה bulk לא נמצא בחשבוניות AMITEA הסוכמות. הזן ידנית."},
    "RAW-MATCHA":             {"price": None,    "note": "אם זה הקוד של מאצ'ה — הזן ידנית."},

    # ── GT Everyday (sugar) — internal/own
    "RAW-SUGAR":              {"price": None,    "note": "סוכר — מסומן GT Everyday. אם רכישה עצמית הזן 0; אחרת הזן עלות פר ק\"ג."},

    # Other
    "RAW-WATER":              {"price": 0.00,    "note": "מים — אספקה פנימית, אין עלות חיצונית."},
}


# ────────────────────────────────────────────────────────────────────────────
# Read filled stock template (cached values written by Excel on save)
# ────────────────────────────────────────────────────────────────────────────
def read_buy_list():
    wb = load_workbook(STOCK_TEMPLATE, data_only=True)
    ws = wb.active

    # Special case: cell with #VALUE! due to non-numeric stock entry — fall back to demand
    SPECIAL_FALLBACK = {
        "RAW-CINNAMON": 26.77,  # demand from earlier; Tom's stock cell had a typo
    }

    items = []
    current_kind = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        cid = ws.cell(row=r, column=2).value
        if a in ("אריזות", "חומרי גלם", "פילטרים וחומרי עזר לייצור"):
            current_kind = a
            continue
        if a == "רכיב":
            continue
        if not cid:
            continue
        name = a
        sup = ws.cell(row=r, column=3).value or "(no supplier)"
        uom = ws.cell(row=r, column=4).value
        dem = ws.cell(row=r, column=5).value
        stk = ws.cell(row=r, column=6).value
        buy = ws.cell(row=r, column=7).value

        # Filter to items that need buying
        if cid in SPECIAL_FALLBACK:
            buy_qty = SPECIAL_FALLBACK[cid]
            note = "מלא ידנית — תא מלאי בטופס המקורי לא ניתן לחישוב"
        else:
            if stk is None:
                # Not filled → skip (Tom didn't tell us to buy)
                continue
            if not isinstance(buy, (int, float)) or buy <= 0:
                continue
            buy_qty = float(buy)
            note = ""

        items.append({
            "kind": current_kind,
            "name": name,
            "cid": cid,
            "supplier": sup,
            "uom": uom or "",
            "buy_qty": buy_qty,
            "note_qty": note,
        })
    return items


# ────────────────────────────────────────────────────────────────────────────
# Quantity rounding policy
# ────────────────────────────────────────────────────────────────────────────
def round_qty(qty, uom):
    """UNIT items round up to integer; weights/volumes to 2 decimals."""
    if uom and uom.upper() == "UNIT":
        return math.ceil(qty)
    return round(qty, 2)


# ────────────────────────────────────────────────────────────────────────────
# Supplier display-name normalization (consolidate variants)
# ────────────────────────────────────────────────────────────────────────────
SUPPLIER_NORMALIZE = {
    "TAVLINEI BAR": "Tavlinei Bar",
}

def norm_supplier(name):
    return SUPPLIER_NORMALIZE.get(name, name)


# ────────────────────────────────────────────────────────────────────────────
# Excel rendering
# ────────────────────────────────────────────────────────────────────────────
HEADER_BG = "FF1F2937"
SECTION_BG = "FFF3F4F6"
INPUT_BG = "FFFFF7CC"
PRICE_FILL_BG = "FFFFF7CC"  # yellow for empty prices to fill
SUBTOTAL_BG = "FFE5E7EB"

TITLE_FONT = Font(name="Arial", size=16, bold=True, color="FFFFFFFF")
SECTION_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFFFF")
COL_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
DEFAULT_FONT = Font(name="Arial", size=10, color="FF111827")
INPUT_FONT = Font(name="Arial", size=10, bold=True, color="FF111827")
SUBTOTAL_FONT = Font(name="Arial", size=10, bold=True, color="FF111827")
SUB_FONT = Font(name="Arial", size=9, italic=True, color="FF6B7280")

THIN = Side(style="thin", color="FFD1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

QTY_FMT_INT = '#,##0;[Red]-#,##0;"-"'
QTY_FMT_DEC = '#,##0.00;[Red]-#,##0.00;"-"'
MONEY_FMT = '#,##0.00" ₪";[Red]-#,##0.00" ₪";"-"'

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def fill_cell(ws, coord, value=None, *, font=None, fill=None, align=None,
              border=BORDER, fmt=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border
    if fmt is not None:
        c.number_format = fmt
    return c


def main():
    items = read_buy_list()
    print(f"Loaded {len(items)} buy-list items from stock template")

    # Group by normalized supplier
    groups = defaultdict(list)
    for it in items:
        sup = norm_supplier(it["supplier"])
        groups[sup].append(it)

    # Sort groups (by supplier name) and items within groups (by name)
    sorted_suppliers = sorted(groups.keys(), key=lambda s: s.lower())
    for sup in sorted_suppliers:
        groups[sup].sort(key=lambda x: (x["name"] or "").lower())

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "תכנית רכש סופית מאי 2026"
    ws.sheet_view.rightToLeft = True

    # Column widths: A name | B code | C supplier | D uom | E qty | F unit price | G line cost | H note
    widths = [38, 26, 24, 8, 14, 16, 18, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.row_dimensions[1].height = 32
    fill_cell(ws, "A1", value="תכנית רכש סופית — מאי 2026",
              font=TITLE_FONT, fill=PatternFill("solid", fgColor=HEADER_BG),
              align=Alignment(horizontal="right", vertical="center"))
    ws.merge_cells("A1:H1")

    fill_cell(ws, "A2",
              value="עמודה F (מחיר ליחידה) — תאים צהובים ריקים: מלא ידנית. עלות שורה, סיכומי ספק וסה\"כ — נוסחאות שמתעדכנות אוטומטית. כל המחירים pre-VAT.",
              font=SUB_FONT, align=ALIGN_RIGHT, border=None)
    ws.merge_cells("A2:H2")

    # Column headers (row 4)
    row = 4
    headers = ["רכיב", "קוד", "ספק", "יח'", "כמות לקנייה", "מחיר ליחידה (₪)", "עלות שורה (₪)", "מקור / הערה"]
    for col, h in zip("ABCDEFGH", headers):
        fill_cell(ws, f"{col}{row}", value=h,
                  font=COL_HEADER_FONT,
                  fill=PatternFill("solid", fgColor=HEADER_BG),
                  align=ALIGN_CENTER)
    row += 1

    grand_total_cells = []  # supplier subtotal cells

    for sup in sorted_suppliers:
        group = groups[sup]
        # supplier section header
        fill_cell(ws, f"A{row}", value=f"ספק: {sup}",
                  font=SECTION_FONT,
                  fill=PatternFill("solid", fgColor=HEADER_BG),
                  align=Alignment(horizontal="right", vertical="center"))
        ws.merge_cells(f"A{row}:H{row}")
        row += 1

        line_cost_cells = []
        for it in group:
            qty = round_qty(it["buy_qty"], it["uom"])
            price_info = PRICES.get(it["cid"], {})
            price = price_info.get("price")
            note_price = price_info.get("note", "")

            fill_cell(ws, f"A{row}", value=it["name"],
                      font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"B{row}", value=it["cid"],
                      font=DEFAULT_FONT, align=ALIGN_CENTER)
            fill_cell(ws, f"C{row}", value=sup,
                      font=DEFAULT_FONT, align=ALIGN_RIGHT)
            fill_cell(ws, f"D{row}", value=it["uom"],
                      font=DEFAULT_FONT, align=ALIGN_CENTER)

            qty_fmt = QTY_FMT_INT if (it["uom"] or "").upper() == "UNIT" else QTY_FMT_DEC
            fill_cell(ws, f"E{row}", value=qty,
                      font=INPUT_FONT,
                      fill=PatternFill("solid", fgColor=INPUT_BG),
                      align=ALIGN_CENTER, fmt=qty_fmt)

            # Price cell: if known → fill in; else leave blank, yellow background
            if price is not None:
                fill_cell(ws, f"F{row}", value=price,
                          font=DEFAULT_FONT, align=ALIGN_CENTER, fmt=MONEY_FMT)
            else:
                fill_cell(ws, f"F{row}", value=None,
                          font=INPUT_FONT,
                          fill=PatternFill("solid", fgColor=PRICE_FILL_BG),
                          align=ALIGN_CENTER, fmt=MONEY_FMT)

            # Line cost = qty * price (Excel treats blank as 0 → cost stays 0 until filled)
            fill_cell(ws, f"G{row}", value=f"=E{row}*F{row}",
                      font=Font(name="Arial", size=10, bold=True, color="FF111827"),
                      align=ALIGN_CENTER, fmt=MONEY_FMT)
            line_cost_cells.append(f"G{row}")

            # Note
            note_text = note_price
            if it["note_qty"]:
                note_text = (note_text + " | " if note_text else "") + it["note_qty"]
            fill_cell(ws, f"H{row}", value=note_text,
                      font=SUB_FONT, align=ALIGN_RIGHT)
            row += 1

        # Supplier subtotal
        subtotal_formula = "=" + "+".join(line_cost_cells) if line_cost_cells else "=0"
        fill_cell(ws, f"A{row}", value=f"סה\"כ {sup}",
                  font=SUBTOTAL_FONT,
                  fill=PatternFill("solid", fgColor=SUBTOTAL_BG),
                  align=ALIGN_RIGHT)
        for col in "BCDEF":
            fill_cell(ws, f"{col}{row}", value="",
                      font=DEFAULT_FONT,
                      fill=PatternFill("solid", fgColor=SUBTOTAL_BG))
        fill_cell(ws, f"G{row}", value=subtotal_formula,
                  font=SUBTOTAL_FONT,
                  fill=PatternFill("solid", fgColor=SUBTOTAL_BG),
                  align=ALIGN_CENTER, fmt=MONEY_FMT)
        fill_cell(ws, f"H{row}", value="",
                  font=DEFAULT_FONT,
                  fill=PatternFill("solid", fgColor=SUBTOTAL_BG))
        grand_total_cells.append(f"G{row}")
        row += 2  # spacer row

    # Grand total
    ws.row_dimensions[row].height = 26
    grand_formula = "=" + "+".join(grand_total_cells) if grand_total_cells else "=0"
    fill_cell(ws, f"A{row}", value="סה\"כ כללי (לפני מע\"מ)",
              font=Font(name="Arial", size=12, bold=True, color="FFFFFFFF"),
              fill=PatternFill("solid", fgColor=HEADER_BG),
              align=Alignment(horizontal="right", vertical="center"))
    for col in "BCDEF":
        fill_cell(ws, f"{col}{row}", value="",
                  font=Font(color="FFFFFFFF"),
                  fill=PatternFill("solid", fgColor=HEADER_BG))
    fill_cell(ws, f"G{row}", value=grand_formula,
              font=Font(name="Arial", size=12, bold=True, color="FFFFFFFF"),
              fill=PatternFill("solid", fgColor=HEADER_BG),
              align=ALIGN_CENTER, fmt=MONEY_FMT)
    fill_cell(ws, f"H{row}", value="",
              fill=PatternFill("solid", fgColor=HEADER_BG))
    row += 1

    # VAT and total-with-VAT
    fill_cell(ws, f"A{row}", value="מע\"מ 18%",
              font=DEFAULT_FONT, align=ALIGN_RIGHT)
    fill_cell(ws, f"G{row}", value=f"=G{row-1}*0.18",
              font=DEFAULT_FONT, align=ALIGN_CENTER, fmt=MONEY_FMT)
    row += 1
    fill_cell(ws, f"A{row}", value="סה\"כ כולל מע\"מ",
              font=Font(name="Arial", size=12, bold=True),
              fill=PatternFill("solid", fgColor=SUBTOTAL_BG),
              align=ALIGN_RIGHT)
    fill_cell(ws, f"G{row}", value=f"=G{row-2}*1.18",
              font=Font(name="Arial", size=12, bold=True),
              fill=PatternFill("solid", fgColor=SUBTOTAL_BG),
              align=ALIGN_CENTER, fmt=MONEY_FMT)

    ws.freeze_panes = "A5"

    print(f"Saving -> {OUT_FILE}")
    try:
        wb.save(OUT_FILE)
        print("Done.")
    except PermissionError:
        print("  (skipped — file is open in Excel; close it and re-run.)")


if __name__ == "__main__":
    main()
